from sheets_auth import get_expense_worksheet, get_income_worksheet, get_subscriptions_worksheet
from openpyxl.utils import column_index_from_string
from datetime import datetime, timedelta
import re
from sheets_config import get_category_columns
from dateutil import parser as dateparser
from rapidfuzz import fuzz
from collections import defaultdict, Counter
import matplotlib.pyplot as plt
import io
import discord
import gspread
from pytz import timezone

# HELPER FUNCTIONS
def _sum_column(ws, col_letter, start_row=3):
    col_idx = column_index_from_string(col_letter)
    values = ws.col_values(col_idx)[start_row - 1:]
    return sum(clean_money(v) for v in values if v.strip())


def clean_money(value: str) -> float:
    """
    Remove $ and , then convert to float.
    Example: '$1,250.00' -> 1250.00
    Empty or invalid strings return 0.0.
    """
    if not value or value.strip() == "":
        return 0.0
    try:
        return float(value.replace('$', '').replace(',', '').strip())
    except Exception as e:
        print(f"[WARN] Failed to clean money value: {value} ({e})")
        return 0.0


def get_local_today():
    return datetime.now(timezone("America/Los_Angeles"))


def find_cell_by_partial_text(ws, text):
    """
    Searches column B for a cell containing `text` (case-insensitive, stripped).
    Returns the cell if found, else None.
    """
    col_b = ws.col_values(2)
    for idx, val in enumerate(col_b, start=1):
        if val and text.lower().strip() in val.lower().strip():
            return ws.cell(idx, 2)
    return None


def extract_amount_from_text(text):
    """
    Extracts the first dollar amount from a string like 'IDEAL = $49.15'
    """
    if not text:
        return 0.0
    match = re.search(r"[-]?\$?([\d,]+(?:\.\d{1,2})?)", text)
    if match:
        try:
            return float(match.group(1).replace(",", ""))
        except Exception as e:
            print(f"[WARN] Failed to parse amount from '{text}': {e}")
            return 0.0
    return 0.0


def resolve_query_persons(discord_user: str, person: str | None) -> list[str]:
    """
    Given discord_user and optional person, return a list of person(s) to query.
    """
    discord_user = (discord_user or "").strip().lower()
    person = (person or "").strip()

    if not person:
        mapping = {
            "hannerish": ["Hannah"],
            ".deebers": ["Brian (BofA)", "Brian (AL)"]
        }
        return mapping.get(discord_user, [])

    if person.lower() == "brian":
        return ["Brian (BofA)", "Brian (AL)"]

    if person in ["Brian (BofA)", "Brian (AL)", "Hannah"]:
        return [person]

    return []  # fallback


# QUERY FUNCTIONS
async def calculate_burn_rate():
    ws = get_income_worksheet()
    try:
        # get all cell values
        all_cells = ws.get_all_values()

        # search manually for a row with 'burn rate' in any cell
        for r, row in enumerate(all_cells, 1):
            for c, cell in enumerate(row, 1):
                if "burn rate" in cell.lower():
                    val_text = cell.strip()
                    print(f"[DEBUG] Found burn rate cell: '{val_text}' at ({r},{c})")

                    parts = val_text.split(":")
                    if len(parts) < 2:
                        print(f"[ERROR] Unexpected format in burn rate cell: {val_text}")
                        return None, None

                    burn_rate_val = parts[1].strip()
                    desc = ""
                    # try to get description 2 columns to the right if it exists
                    if c + 2 <= len(row):
                        desc = row[c + 1].strip()
                    return burn_rate_val, desc

        print("[ERROR] Could not find any cell containing 'burn rate'")
        return None, None

    except Exception as e:
        print(f"[ERROR] Failed to fetch burn rate: {e}")
        return None, None


async def check_rent_paid():
    ws = get_income_worksheet()
    try:
        cell = ws.find("Rent")
        amount_cell = ws.cell(cell.row, cell.col + 1).value
        if amount_cell:
            cleaned = clean_money(amount_cell)
            if cleaned > 0:
                return True, cleaned
    except Exception as e:
        print(f"[ERROR] Failed to check rent paid: {e}")
    return False, 0.0


async def check_smud_paid():
    ws = get_income_worksheet()
    try:
        cell = ws.find("SMUD")
        amount_cell = ws.cell(cell.row, cell.col + 1).value
        if amount_cell:
            cleaned = clean_money(amount_cell)
            if cleaned > 0:
                return True, cleaned
    except Exception as e:
        print(f"[ERROR] Failed to check SMUD paid: {e}")
    return False, 0.0

async def check_student_loan_paid():
    ws = get_income_worksheet()
    try:
        cell = ws.find("Student Loan Payment")
        amount_cell = ws.cell(cell.row, cell.col + 1).value
        if amount_cell:
            cleaned = clean_money(amount_cell)
            if cleaned > 0:
                return True, cleaned
    except Exception as e:
        print(f"[ERROR] Failed to check student loan paid: {e}")
    return False, 0.0


async def total_spent_at_store(store, persons, top_n=5):
    ws = get_expense_worksheet()
    today = get_local_today()
    total = 0.0
    matches = []

    store_norm = store.lower().replace(" ", "")
    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        location_col_letter = config["columns"].get("location")
        person_col_letter = config["columns"].get("person")

        if not location_col_letter or not person_col_letter:
            continue  # skip if either missing

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        location_idx = column_index_from_string(location_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, location_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            location_str = row[location_idx].strip().lower().replace(" ", "")
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not location_str or not person_str:
                continue

            if person_str not in persons:
                continue  # skip if this person isn’t one of the queried

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    similarity = fuzz.partial_ratio(store_norm, location_str)
                    if similarity >= 80:
                        amt = clean_money(amount_str)
                        total += amt
                        matches.append((
                            date_obj,
                            row[location_idx],
                            amt,
                            category
                        ))
                        print(f"[MATCH] {date_obj.date()} | {row[location_idx]} | ${amt:.2f} | sim: {similarity}% | {person_str}")
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    matches.sort(key=lambda x: x[0], reverse=True)

    return total, matches[:top_n]


async def highest_expense_category(persons):
    ws = get_expense_worksheet()
    category_totals = {}
    today = get_local_today()

    categories = {
        'grocery': {'amount': 'B', 'person': 'D'},
        'gas': {'amount': 'I', 'person': 'J'},
        'food': {'amount': 'P', 'person': 'R'},
        'shopping': {'amount': 'X', 'person': 'Z'}
    }

    for category, cols in categories.items():
        amount_col_idx = column_index_from_string(cols['amount']) - 1
        person_col_idx = column_index_from_string(cols['person']) - 1

        rows = ws.get_all_values()[2:]  # skip header

        total = 0.0
        for row in rows:
            if max(amount_col_idx, person_col_idx) >= len(row):
                continue

            amount_str = row[amount_col_idx].strip()
            person_str = row[person_col_idx].strip()

            if not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                amt = clean_money(amount_str)
                total += amt
            except Exception as e:
                print(f"[WARN] Failed to parse row: {e}")
                continue

        category_totals[category] = total

    if not category_totals:
        return None, 0.0

    highest = max(category_totals.items(), key=lambda x: x[1])
    return highest  # (category, amount)


async def total_income():
    ws = get_income_worksheet()
    try:
        cell = ws.find("Monthly Income:")
        income_val = ws.cell(cell.row, cell.col + 1).value
        income_val = income_val.strip()

        # Try to convert to float if possible
        return clean_money(income_val)
    except (AttributeError, ValueError) as e:
        print(f"[WARN] Income value is missing or invalid: {e}")
        return 0.0
    except Exception as e:
        print(f"[ERROR] Failed to fetch income: {e}")
        return 0.0


async def remaining_budget():
    ws = get_income_worksheet()
    try:
        cell = ws.find("Margins:")
        val = ws.cell(cell.row, cell.col + 2).value
        remaining_budget = clean_money(val)
        return remaining_budget
    except Exception as e:
        print(f"[ERROR] Failed to get remaining budget: {e}")
        return 0.0


async def average_daily_spend(persons):
    ws = get_expense_worksheet()
    try:
        today = get_local_today()
        day_of_month = today.day

        shopping_total = 0.0
        food_total = 0.0

        rows = ws.get_all_values()[2:]  # skip header rows

        # Indices for Shopping & Food
        shop_date_idx = column_index_from_string('V') - 1
        shop_amount_idx = column_index_from_string('X') - 1
        shop_person_idx = column_index_from_string('Z') - 1

        food_date_idx = column_index_from_string('N') - 1
        food_amount_idx = column_index_from_string('P') - 1
        food_person_idx = column_index_from_string('R') - 1

        for row in rows:
            # Shopping
            if len(row) > max(shop_date_idx, shop_amount_idx, shop_person_idx):
                date_str = row[shop_date_idx].strip()
                amount_str = row[shop_amount_idx].strip()
                person_str = row[shop_person_idx].strip()

                if date_str and amount_str and person_str in persons:
                    date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                    if date_obj.month == today.month and date_obj.year == today.year:
                        amt = clean_money(amount_str)
                        shopping_total += amt

            # Food
            if len(row) > max(food_date_idx, food_amount_idx, food_person_idx):
                date_str = row[food_date_idx].strip()
                amount_str = row[food_amount_idx].strip()
                person_str = row[food_person_idx].strip()

                if date_str and amount_str and person_str in persons:
                    date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                    if date_obj.month == today.month and date_obj.year == today.year:
                        amt = clean_money(amount_str)
                        food_total += amt

        total_spent = shopping_total + food_total
        avg_daily_spend = total_spent / day_of_month if day_of_month else 0.0
        return round(avg_daily_spend, 2)

    except Exception as e:
        print(f"[ERROR] Failed to compute average daily spend: {e}")
        return None


async def expense_breakdown_percentages(persons: list[str]):
    ws = get_expense_worksheet()
    category_amounts = {'grocery': 0.0, 'gas': 0.0, 'food': 0.0, 'shopping': 0.0}
    grand_total = 0.0

    # per-person row and total cell maps
    person_row_map = {
        "Brian (AL)": 3,
        "Brian (BofA)": 4,
        "Hannah": 5
    }

    total_cell_map = {
        "Brian (AL)": "AE27",
        "Brian (BofA)": "AE28",
        "Hannah": "AE31"
    }

    category_cells = {
        'grocery': 'F{}',
        'gas': 'L{}',
        'food': 'T{}',
        'shopping': 'AB{}'
    }

    # sum categories & total over all persons
    for person in persons:
        row = person_row_map.get(person)
        total_cell = total_cell_map.get(person)

        if not row or not total_cell:
            print(f"[ERROR] Unknown person: {person}")
            continue

        total_val = ws.acell(total_cell).value
        grand_total += clean_money(total_val)

        for category, cell_fmt in category_cells.items():
            cell_ref = cell_fmt.format(row)
            c_row = int(re.sub(r'\D', '', cell_ref))
            c_col = column_index_from_string(re.sub(r'\d', '', cell_ref))
            val = ws.cell(c_row, c_col).value
            amount = clean_money(val)
            category_amounts[category] += amount

    if grand_total == 0:
        print(f"[WARN] Combined total for {persons} is 0. Cannot calculate breakdown.")
        return {}

    # build result
    breakdown = {}
    for cat, amt in category_amounts.items():
        pct = round(amt / grand_total * 100, 2) if grand_total > 0 else 0
        breakdown[cat] = {
            "amount": round(amt, 2),
            "percentage": pct
        }

    result = {
        "categories": breakdown,
        "grand_total": round(grand_total, 2)
    }

    return result


async def total_for_category(category, persons):
    ws = get_expense_worksheet()
    today = get_local_today()

    category = category.lower()
    config = {
        'grocery': {'amount': 'B', 'person': 'D', 'start_row': 3},
        'gas': {'amount': 'I', 'person': 'J', 'start_row': 3},
        'food': {'amount': 'P', 'person': 'R', 'start_row': 3},
        'shopping': {'amount': 'X', 'person': 'Z', 'start_row': 3}
    }

    if category not in config:
        print(f"[ERROR] Unknown category: {category}")
        return 0.0

    cols = config[category]
    amount_idx = column_index_from_string(cols['amount']) - 1
    person_idx = column_index_from_string(cols['person']) - 1

    rows = ws.get_all_values()[cols['start_row'] - 1:]
    total = 0.0

    for row in rows:
        if max(amount_idx, person_idx) >= len(row):
            continue

        amount_str = row[amount_idx].strip()
        person_str = row[person_idx].strip()

        if not amount_str or not person_str:
            continue

        if person_str not in persons:
            continue

        try:
            amt = clean_money(amount_str)
            total += amt
        except Exception as e:
            print(f"[WARN] Failed to parse row: {e}")
            continue

    return round(total, 2)


async def largest_single_expense(persons):
    ws = get_expense_worksheet()
    rows = ws.get_all_values()[2:]  # skip header rows

    configs = {
        "food": {
            "date": "N",
            "item": "O",
            "amount": "P",
            "location": "Q",
            "person": "R"
        },
        "shopping": {
            "date": "V",
            "item": "W",
            "amount": "X",
            "location": "Y",
            "person": "Z"
        }
    }

    max_amount = 0.0
    result = None

    for row in rows:
        for category, cols in configs.items():
            try:
                date_idx = column_index_from_string(cols["date"]) - 1
                item_idx = column_index_from_string(cols["item"]) - 1
                amount_idx = column_index_from_string(cols["amount"]) - 1
                location_idx = column_index_from_string(cols["location"]) - 1
                person_idx = column_index_from_string(cols["person"]) - 1

                if max(date_idx, item_idx, amount_idx, location_idx, person_idx) >= len(row):
                    continue

                person_str = row[person_idx].strip()
                if person_str not in persons:
                    continue

                amount_str = row[amount_idx].strip()
                if not amount_str:
                    continue

                amt = clean_money(amount_str)

                if amt > max_amount:
                    max_amount = amt
                    result = {
                        "category": category,
                        "amount": round(amt, 2),
                        "date": row[date_idx],
                        "item": row[item_idx],
                        "location": row[location_idx]
                    }

            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    return result


async def top_n_expenses_all_categories(persons, n=5):
    ws = get_expense_worksheet()
    rows = ws.get_all_values()[2:]  # skip header

    configs = {
        "grocery": {
            "date": "A", "item": None, "amount": "B", "location": "C", "person": "D"
        },
        "gas": {
            "date": "H", "item": None, "amount": "I", "location": None, "person": "J"
        },
        "food": {
            "date": "N", "item": "O", "amount": "P", "location": "Q", "person": "R"
        },
        "shopping": {
            "date": "V", "item": "W", "amount": "X", "location": "Y", "person": "Z"
        }
    }

    expenses = []

    for row in rows:
        for category, cols in configs.items():
            try:
                date_idx = column_index_from_string(cols["date"]) - 1
                amount_idx = column_index_from_string(cols["amount"]) - 1
                person_idx = column_index_from_string(cols["person"]) - 1

                item_idx = column_index_from_string(cols["item"]) - 1 if cols["item"] else None
                location_idx = column_index_from_string(cols["location"]) - 1 if cols["location"] else None

                if max(date_idx, amount_idx, person_idx, *(i for i in [item_idx, location_idx] if i is not None)) >= len(row):
                    continue

                person_str = row[person_idx].strip()
                if person_str not in persons:
                    continue

                amount_str = row[amount_idx].strip()
                if not amount_str:
                    continue

                amt = clean_money(amount_str)
                if amt <= 0:
                    continue

                expense = {
                    "category": category,
                    "amount": round(amt, 2),
                    "date": row[date_idx],
                    "item": row[item_idx] if item_idx is not None and item_idx < len(row) else "",
                    "location": row[location_idx] if location_idx is not None and location_idx < len(row) else ""
                }
                expenses.append(expense)

            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    # Sort by amount descending
    expenses.sort(key=lambda x: x["amount"], reverse=True)

    return expenses[:n]


async def spent_this_week(persons):
    ws = get_expense_worksheet()
    today = get_local_today()
    start_of_week = today - timedelta(days=today.weekday())  # Monday
    total = 0.0

    category_columns = get_category_columns  # dict

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        person_col_letter = config["columns"].get("person")

        if not person_col_letter:
            continue  # skip categories that don’t track person

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = dateparser.parse(date_str, dayfirst=False, yearfirst=False)
                if date_obj.date() >= start_of_week.date():
                    amt = clean_money(amount_str)
                    print(f"[MATCH] {date_obj.date()} ${amt:.2f} ({person_str})")
                    total += amt
                else:
                    print(f"[SKIP] {date_obj.date()} before start of week {start_of_week.date()}")
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    return round(total, 2)



async def projected_spending(persons):
    today = get_local_today()
    ws = get_expense_worksheet()
    total_so_far = 0.0

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        person_col_letter = config["columns"].get("person")

        if not person_col_letter:
            continue  # skip categories without a person column

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    amt = clean_money(amount_str)
                    total_so_far += amt
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    # Project to end of month
    if today.month == 12:
        next_month = datetime(today.year + 1, 1, 1)
    else:
        next_month = datetime(today.year, today.month + 1, 1)

    days_in_month = (next_month - timedelta(days=1)).day
    days_passed = today.day
    daily_avg = total_so_far / days_passed if days_passed else 0.0
    projected = daily_avg * days_in_month

    return round(projected, 2)


async def weekend_vs_weekday(persons):
    ws = get_expense_worksheet()
    weekend = 0.0
    weekday = 0.0

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        person_col_letter = config["columns"].get("person")

        if not person_col_letter:
            continue  # skip categories without a person column

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                amt = clean_money(amount_str)

                if date_obj.weekday() >= 5:  # Saturday or Sunday
                    weekend += amt
                else:
                    weekday += amt

            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    return round(weekend, 2), round(weekday, 2)


async def no_spend_days(persons):
    ws = get_expense_worksheet()
    today = get_local_today()
    days_with_expense = set()

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        person_col_letter = config["columns"].get("person")

        if not person_col_letter:
            continue  # skip if no person column

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    days_with_expense.add(date_obj.day)
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    all_days = set(range(1, today.day + 1))
    no_spend = sorted(all_days - days_with_expense)
    return len(no_spend), no_spend


async def total_spent_on_item(item, persons, top_n=5):
    ws = get_expense_worksheet()
    today = get_local_today()
    total = 0.0
    matches = []

    item_norm = item.lower().replace(" ", "")

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        item_col_letter = config["columns"].get("item")
        person_col_letter = config["columns"].get("person")

        if not item_col_letter or not person_col_letter:
            continue  # skip if either is missing

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        item_idx = column_index_from_string(item_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, item_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            item_str = row[item_idx].strip().lower().replace(" ", "")
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not item_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    similarity = fuzz.partial_ratio(item_norm, item_str)
                    if similarity >= 80:
                        amt = clean_money(amount_str)
                        total += amt
                        matches.append((
                            date_obj,
                            row[item_idx],
                            amt,
                            category,
                            person_str
                        ))
                        print(f"[MATCH] {date_obj.date()} | {row[item_idx]} | ${amt:.2f} | sim: {similarity}% | {person_str}")
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    matches.sort(key=lambda x: x[0], reverse=True)

    return total, matches[:top_n]


async def daily_spending_calendar(persons):
    ws = get_expense_worksheet()
    today = get_local_today()
    daily_totals = defaultdict(float)

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        person_col_letter = config["columns"].get("person")

    # skip categories without person
        if not person_col_letter:
            continue

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        person_idx = column_index_from_string(person_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, person_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            person_str = row[person_idx].strip()

            if not date_str or not amount_str or not person_str:
                continue

            if person_str not in persons:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    amt = clean_money(amount_str)
                    daily_totals[date_obj.day] += amt
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    # Fill missing days with 0
    for day in range(1, today.day + 1):
        daily_totals.setdefault(day, 0.0)

    # Prepare sorted list
    sorted_days = sorted(daily_totals.items())

    # Text summary
    text_lines = [f"{today.strftime('%B %Y')} Daily Spending:"]
    for day, amt in sorted_days:
        text_lines.append(f"{day:02d}: ${amt:.2f}")
    text_summary = "\n".join(text_lines)

    # Plot
    days = [day for day, _ in sorted_days]
    amounts = [amt for _, amt in sorted_days]

    fig, ax = plt.subplots(figsize=(10, 5))
    ax.bar(days, amounts, color='skyblue')
    ax.set_xlabel("Day of Month")
    ax.set_ylabel("Amount Spent ($)")
    ax.set_title(f"Daily Spending — {today.strftime('%B %Y')}")
    ax.set_xticks(days)
    plt.tight_layout()

    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=150)
    buf.seek(0)
    plt.close(fig)

    chart_file = discord.File(fp=buf, filename="daily_spending_calendar.png")

    return text_summary, chart_file


async def best_worst_day_of_week():
    ws = get_expense_worksheet()
    today = get_local_today()
    weekday_totals = defaultdict(float)
    weekday_counts = defaultdict(int)

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()

            if not date_str or not amount_str:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    weekday = date_obj.weekday()  # 0=Monday, 6=Sunday
                    amt = clean_money(amount_str)
                    weekday_totals[weekday] += amt
                    weekday_counts[weekday] += 1
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    averages = {}
    for wd in range(7):
        if weekday_counts[wd] > 0:
            averages[wd] = weekday_totals[wd] / weekday_counts[wd]
        else:
            averages[wd] = 0.0

    best_day = min(averages.items(), key=lambda x: x[1])
    worst_day = max(averages.items(), key=lambda x: x[1])

    # Map weekday numbers to names
    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]

    return {
        "best": (weekday_names[best_day[0]], round(best_day[1], 2)),
        "worst": (weekday_names[worst_day[0]], round(worst_day[1], 2)),
    }


async def longest_no_spend_streak():
    ws = get_expense_worksheet()
    today = get_local_today()
    daily_totals = {day: 0.0 for day in range(1, today.day + 1)}

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()

            if not date_str or not amount_str:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    amt = clean_money(amount_str)
                    daily_totals[date_obj.day] += amt
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    # Find longest streak
    longest = 0
    current = 0
    start = None
    best_start = None
    best_end = None

    for day in range(1, today.day + 1):
        if daily_totals[day] == 0.0:
            if current == 0:
                start = day
            current += 1
            if current > longest:
                longest = current
                best_start = start
                best_end = day
        else:
            current = 0

    if longest == 0:
        return None  # no streak found

    # Return as a tuple: (length, start_day, end_day)
    return (longest, best_start, best_end)


async def days_budget_lasts():
    # Reuse your existing helpers
    remaining = await remaining_budget()
    avg_daily = await average_daily_spend()

    if avg_daily == 0:
        return None  # avoid division by 0

    estimated_days = remaining / avg_daily

    return max(0, round(estimated_days, 1))  # never negative


async def most_frequent_purchases(n=3):
    ws = get_expense_worksheet()
    today = get_local_today()
    item_counts = Counter()
    item_totals = defaultdict(float)

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        item_col_letter = config["columns"].get("item")

        if not item_col_letter:
            continue  # skip if no item column

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        item_idx = column_index_from_string(item_col_letter) - 1

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx, item_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            item_str = row[item_idx].strip().lower()

            if not date_str or not amount_str or not item_str:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.month == today.month and date_obj.year == today.year:
                    amt = clean_money(amount_str)
                    item_counts[item_str] += 1
                    item_totals[item_str] += amt
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    if not item_counts:
        return []

    most_common = item_counts.most_common(n)
    result = []
    for item_name, count in most_common:
        result.append({
            "item": item_name,
            "count": count,
            "total": round(item_totals[item_name], 2)
        })

    return result


async def expenses_on_day(day_str):
    """
    Find all expenses on a specific day.
    Supports: MM/DD, MM/DD/YYYY, YYYY-MM-DD, or natural language dates.
    """
    try:
        # Robustly parse the input date string
        target_date = dateparser.parse(day_str)
    except Exception as e:
        print(f"[ERROR] Failed to parse date: {day_str} — {e}")
        return None, None

    ws = get_expense_worksheet()
    entries = []
    total = 0.0

    category_columns = get_category_columns

    for category, config in category_columns.items():
        start_row = config["start_row"]
        date_col_letter = config["columns"]["date"]
        amount_col_letter = config["columns"]["amount"]
        item_col_letter = config["columns"].get("item")
        location_col_letter = config["columns"].get("location")

        date_idx = column_index_from_string(date_col_letter) - 1
        amount_idx = column_index_from_string(amount_col_letter) - 1
        item_idx = column_index_from_string(item_col_letter) - 1 if item_col_letter else None
        location_idx = column_index_from_string(location_col_letter) - 1 if location_col_letter else None

        rows = ws.get_all_values()[start_row - 1:]

        for row in rows:
            if max(date_idx, amount_idx) >= len(row):
                continue

            date_str = row[date_idx].strip()
            amount_str = row[amount_idx].strip()
            item_str = row[item_idx].strip() if item_idx and len(row) > item_idx else ""
            location_str = row[location_idx].strip() if location_idx and len(row) > location_idx else ""

            if not date_str or not amount_str:
                continue

            try:
                date_obj = datetime.strptime(date_str, "%m/%d/%Y")
                if date_obj.date() == target_date.date():
                    amt = clean_money(amount_str)
                    total += amt
                    entries.append({
                        "category": category,
                        "item": item_str,
                        "location": location_str,
                        "amount": round(amt, 2)
                    })
            except Exception as e:
                print(f"[WARN] Skipping row: {e}")
                continue

    return entries, round(total, 2)


async def list_subscriptions():
    ws = get_subscriptions_worksheet()
    needs = []
    wants = []
    needs_total = 0.0
    wants_total = 0.0

    rows = ws.get_all_values()

    # Needs and Wants start at row 7 (0-based index = 6)
    for row in rows[6:]:
        # Needs
        name_needs = row[1].strip() if len(row) > 1 else ""
        amount_needs = row[2].strip() if len(row) > 2 else ""
        if name_needs and amount_needs:
            try:
                amt = float(amount_needs.replace("$", ""))
                needs.append((name_needs, amt))
                needs_total += amt
            except:
                pass

        # Wants
        name_wants = row[4].strip() if len(row) > 4 else ""
        amount_wants = row[5].strip() if len(row) > 5 else ""
        if name_wants and amount_wants:
            try:
                amt = float(amount_wants.replace("$", ""))
                wants.append((name_wants, amt))
                wants_total += amt
            except:
                pass

    return needs, round(needs_total, 2), wants, round(wants_total, 2)


def log_payment(category_label, amount):
    """
    Finds the row where column B matches category_label (case-insensitive),
    and writes the amount to column C of that row.
    """
    ws = get_income_worksheet()
    rows = ws.get_all_values()

    for row_idx, row in enumerate(rows):
        if len(row) < 2:
            continue

        label_cell = row[1].strip().lower()  # Column B
        if label_cell.startswith(category_label.lower()):
            # Write to column C (3)
            cell_to_update = gspread.utils.rowcol_to_a1(row_idx + 1, 3)
            ws.update_acell(cell_to_update, str(amount))
            print(f"[INFO] Logged ${amount} for {category_label} at {cell_to_update}")
            return True

    print(f"[ERROR] Could not find category '{category_label}' in income sheet.")
    return False


def log_rent_paid(amount):
    return log_payment("rent", amount)


def log_smud_paid(amount):
    return log_payment("smud", amount)


def log_student_loan_paid(amount):
    return log_payment("student loan payment", amount)


async def check_1st_savings_deposited():
    ws = get_income_worksheet()
    try:
        # robust text match
        cell = find_cell_by_partial_text(ws, "Enter 1st Paycheck Deposit")
        if not cell:
            raise ValueError("Could not find 'Enter 1st Paycheck Deposit'")

        row, col = cell.row, cell.col

        # actual = 3 right
        actual_val = ws.cell(row, col + 3).value
        actual_amount = clean_money(actual_val)

        # ideal = 1 right
        ideal_val = ws.cell(row, col + 1).value
        ideal_amount = extract_amount_from_text(ideal_val)

        # minimum = 1 right & 1 down
        min_val = ws.cell(row + 1, col + 1).value
        minimum_amount = extract_amount_from_text(min_val)

        deposited = actual_amount > 0

        return {
            "deposited": deposited,
            "actual": actual_amount,
            "ideal": ideal_amount,
            "minimum": minimum_amount
        }

    except Exception as e:
        print(f"[ERROR] Failed to check 1st savings deposited: {e}")
        return {
            "deposited": False,
            "actual": 0.0,
            "ideal": 0.0,
            "minimum": 0.0
        }


async def check_2nd_savings_deposited():
    ws = get_income_worksheet()
    try:
        cell = find_cell_by_partial_text(ws, "Enter 2nd Paycheck Deposit")
        if not cell:
            raise ValueError("Could not find 'Enter 2nd Paycheck Deposit'")

        row, col = cell.row, cell.col

        # actual = 3 right
        actual_val = ws.cell(row, col + 3).value
        actual_amount = clean_money(actual_val)

        # ideal = 1 right & 1 up
        ideal_val = ws.cell(row - 1, col + 1).value
        ideal_amount = extract_amount_from_text(ideal_val)

        # minimum = 1 right
        min_val = ws.cell(row, col + 1).value
        minimum_amount = extract_amount_from_text(min_val)

        deposited = actual_amount > 0

        return {
            "deposited": deposited,
            "actual": actual_amount,
            "ideal": ideal_amount,
            "minimum": minimum_amount
        }

    except Exception as e:
        print(f"[ERROR] Failed to check 2nd savings deposited: {e}")
        return {
            "deposited": False,
            "actual": 0.0,
            "ideal": 0.0,
            "minimum": 0.0
        }


def log_1st_savings(amount):
    """
    Logs the 1st savings deposit amount by writing it 3 columns to the right
    of the cell containing 'Enter 1st Paycheck Deposit'.
    """
    ws = get_income_worksheet()
    try:
        cell = find_cell_by_partial_text(ws, "Enter 1st Paycheck Deposit")
        if not cell:
            raise ValueError("Could not find 'Enter 1st Paycheck Deposit'")

        row, col = cell.row, cell.col
        target_cell = gspread.utils.rowcol_to_a1(row, col + 3)

        ws.update_acell(target_cell, str(amount))
        print(f"[INFO] Logged 1st savings deposit: ${amount} at {target_cell}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to log 1st savings deposit: {e}")
        return False


def log_2nd_savings(amount):
    """
    Logs the 2nd savings deposit amount by writing it 3 columns to the right
    of the cell containing 'Enter 2nd Paycheck Deposit'.
    """
    ws = get_income_worksheet()
    try:
        cell = find_cell_by_partial_text(ws, "Enter 2nd Paycheck Deposit")
        if not cell:
            raise ValueError("Could not find 'Enter 2nd Paycheck Deposit'")

        row, col = cell.row, cell.col
        target_cell = gspread.utils.rowcol_to_a1(row, col + 3)

        ws.update_acell(target_cell, str(amount))
        print(f"[INFO] Logged 2nd savings deposit: ${amount} at {target_cell}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to log 2nd savings deposit: {e}")
        return False


def log_need_expense(description, amount):
    """
    Logs a Need expense by inserting a row above the <Enter Transaction> marker
    in the Needs section of the income sheet.
    Writes description in column B and amount in column C.
    """
    ws = get_income_worksheet()
    try:
        # find the <Enter Transaction> marker
        cell = ws.find("<Enter Transaction>")
        insert_row_idx = cell.row

        # insert a blank row above
        ws.insert_row([], index=insert_row_idx)

        # write description and amount
        ws.update_acell(f"B{insert_row_idx}", description)
        ws.update_acell(f"C{insert_row_idx}", str(amount))

        print(f"[INFO] Logged Need expense: '{description}' - ${amount} at row {insert_row_idx}")
        return True

    except Exception as e:
        print(f"[ERROR] Failed to log Need expense: {e}")
        return False
