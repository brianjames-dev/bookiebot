# expense & income logging

import copy
from datetime import datetime
import logging
from typing import Any, Literal, cast, overload
from openpyxl.utils import column_index_from_string, get_column_letter
from bookiebot.ui.card import CardButtonView
import asyncio
import os
from zoneinfo import ZoneInfo
from bookiebot.sheets.config import expense_category_label, get_category_columns, normalize_expense_category
from bookiebot.sheets.utils import resolve_query_persons
from bookiebot.sheets.repo import get_sheets_repo
from bookiebot.sheets.routing import get_current_discord_user_id
from bookiebot.sheets.undo import UndoAction, _sheet_user_entered_value, _update_contiguous_row, record_undo_action

logger = logging.getLogger(__name__)

_INCOME_SOURCE_PLACEHOLDERS = {"<enter employer>", "<enter source>"}


# Temporary memory for user interactions (used for dropdown callbacks)
pending_data_by_user = {}


def _logged_expense_label(category: str) -> str:
    if category == "need_expenses":
        return "Need expense"
    return f"{expense_category_label(category)} expense"


async def _expense_sheet_with_retry(attempts: int = 2):
    last_error: Exception | None = None
    for attempt in range(max(1, attempts)):
        try:
            return get_sheets_repo().expense_sheet()
        except Exception as exc:
            last_error = exc
            if attempt < attempts - 1:
                await asyncio.sleep(0.75)
    assert last_error is not None
    raise last_error


async def write_to_sheet(data, message):
    if data["type"] == "income":
        await write_income_to_sheet(data, message)
        return

    await write_expense_to_sheet(data, message)


async def write_income_to_sheet(data, message):
    try:
        ws = get_sheets_repo().income_sheet()
    except Exception as e:
        logger.error("Could not access income sheet", extra={"exception": str(e)})
        if message:
            await message.channel.send("Error accessing income sheet.")
        return
    if ws is None:
        if message:
            await message.channel.send("Error accessing income sheet.")
        return

    try:
        row, description, amount = cast(
            tuple[int, str, Any],
            log_income_row(data, ws, return_action_id=False),
        )
    except Exception as e:
        logger.error("Could not log income", extra={"exception": str(e)})
        return

    logger.info("Income logged", extra={"description": description, "amount": amount, "row": row})
    if message:
        await message.channel.send(
            f"Income logged: ${amount} from {data.get('source')}"
        )


@overload
def log_income_row(
    data: dict[str, Any],
    worksheet: Any,
    *,
    return_action_id: Literal[True],
    metadata_extra: dict | None = None,
) -> tuple[int, str, Any, str | None]:
    ...


@overload
def log_income_row(
    data: dict[str, Any],
    worksheet: Any,
    *,
    return_action_id: Literal[False] = False,
    metadata_extra: dict | None = None,
) -> tuple[int, str, Any]:
    ...


def log_income_row(data: dict[str, Any], worksheet: Any, *, return_action_id: bool = False, metadata_extra: dict | None = None):
    try:
        summary_cell = worksheet.find("Monthly Income:")
        summary_row = summary_cell.row
    except Exception as e:
        logger.error("Could not find 'Monthly Income:' in the sheet.", extra={"exception": str(e)})
        raise

    layout = _income_sheet_layout(worksheet, summary_row)
    source_column = layout["source"]
    amount_column = layout["amount"]
    date_column = layout.get("date")

    description = f"{data.get('source', '')} {data.get('label', '')}".strip()
    amount = data.get("amount", "")
    row_values: list[Any] = [""] * max(source_column, amount_column, date_column or 0)
    row_values[source_column - 1] = description
    row_values[amount_column - 1] = amount
    if date_column:
        row_values[date_column - 1] = _income_date_text(data.get("date"))

    rows = worksheet.get_all_values()
    placeholder_rows = _trailing_income_placeholder_rows(
        rows,
        layout=layout,
        summary_row=summary_row,
    )
    if placeholder_rows:
        income_row = placeholder_rows[0]
        first_income_column = min(source_column, amount_column, date_column or source_column)
        _update_contiguous_row(
            worksheet,
            income_row,
            list(range(first_income_column, len(row_values) + 1)),
            row_values[first_income_column - 1 :],
        )
        needs_placeholder = len(placeholder_rows) == 1
    else:
        income_row = summary_row
        worksheet.insert_row(
            row_values,
            index=income_row,
            value_input_option="USER_ENTERED",
            inherit_from_before=income_row > 1,
        )
        if income_row - 1 > layout["header_row"]:
            _copy_income_row_properties(
                worksheet,
                source_row=income_row - 1,
                target_row=income_row,
                start_column=min(source_column, amount_column, date_column or source_column),
                end_column=len(row_values),
            )
        needs_placeholder = True

    if needs_placeholder:
        worksheet.insert_row(
            _income_placeholder_values(layout),
            index=income_row + 1,
            value_input_option="USER_ENTERED",
            inherit_from_before=True,
        )
        _copy_income_row_properties(
            worksheet,
            source_row=income_row,
            target_row=income_row + 1,
            start_column=min(source_column, amount_column, date_column or source_column),
            end_column=max(source_column, amount_column, date_column or source_column),
        )

    _repair_income_summary_formula(worksheet, layout)
    layout_metadata = {
        "income_source_column": str(source_column),
        "income_amount_column": str(amount_column),
    }
    if date_column:
        layout_metadata["income_date_column"] = str(date_column)
    action_id = record_undo_action(
        get_current_discord_user_id(),
        UndoAction(
            worksheet="income",
            kind="delete_row",
            row=income_row,
            columns=[],
            previous_values=[],
            new_values=[str(value) for value in row_values],
            metadata={
                "type": "income",
                "source": str(data.get("source") or ""),
                **(metadata_extra or {}),
                **layout_metadata,
            },
            description=f"income ${amount} from {data.get('source')}",
        ),
    )
    if return_action_id:
        return income_row, description, amount, action_id
    return income_row, description, amount


def _income_sheet_layout(worksheet: Any, summary_row: int) -> dict[str, Any]:
    rows = worksheet.get_all_values()
    for row_number, row in enumerate(rows[: max(0, summary_row - 1)], start=1):
        columns_by_header: dict[str, int] = {}
        source_placeholder = "<Enter Source>"
        for column, value in enumerate(row, start=1):
            normalized = str(value).strip().rstrip(":").strip().lower()
            if normalized in {"date", "amount"}:
                columns_by_header[normalized] = column
            elif normalized in {"source", "employer"}:
                columns_by_header["source"] = column
                source_placeholder = "<Enter Employer>" if normalized == "employer" else "<Enter Source>"
        if "source" in columns_by_header and "amount" in columns_by_header:
            return {
                **columns_by_header,
                "header_row": row_number,
                "source_placeholder": source_placeholder,
            }

    return {
        "source": 2,
        "amount": 3,
        "header_row": 1,
        "source_placeholder": "<Enter Employer>",
    }


def _trailing_income_placeholder_rows(
    rows: list[list[str]],
    *,
    layout: dict[str, Any],
    summary_row: int,
) -> list[int]:
    source_column = layout["source"]
    amount_column = layout["amount"]
    first_income_row = layout["header_row"] + 1
    trailing: list[int] = []
    for row_number in range(summary_row - 1, first_income_row - 1, -1):
        row = rows[row_number - 1] if row_number <= len(rows) else []
        source = row[source_column - 1] if source_column <= len(row) else ""
        amount = row[amount_column - 1] if amount_column <= len(row) else ""
        if not _is_income_placeholder_row(source, amount):
            break
        trailing.append(row_number)
    return sorted(trailing)


def _is_income_placeholder_row(source: Any, amount: Any) -> bool:
    normalized_source = str(source or "").strip().lower()
    if normalized_source not in _INCOME_SOURCE_PLACEHOLDERS:
        return False
    normalized_amount = str(amount or "").strip().replace("$", "").replace(",", "")
    if normalized_amount == "":
        return True
    try:
        return float(normalized_amount) == 0
    except ValueError:
        return False


def _income_placeholder_values(layout: dict[str, Any]) -> list[Any]:
    width = max(layout["source"], layout["amount"], layout.get("date") or 0)
    values: list[Any] = [""] * width
    values[layout["source"] - 1] = layout["source_placeholder"]
    values[layout["amount"] - 1] = 0
    return values


def _copy_income_row_properties(
    worksheet: Any,
    *,
    source_row: int,
    target_row: int,
    start_column: int,
    end_column: int,
) -> None:
    """Copy the income seed row's cell properties after a Sheets row insertion.

    Google Sheets' ``inheritFromBefore`` copies most formatting, but live Sheets
    testing showed that it omits notes and some border details. Reapplying the
    source row's explicit cell properties keeps every generated income row
    consistent with the template without copying its values.
    """
    spreadsheet = getattr(worksheet, "spreadsheet", None)
    if spreadsheet is None or not hasattr(spreadsheet, "fetch_sheet_metadata"):
        return
    if not hasattr(spreadsheet, "batch_update"):
        return

    try:
        title = str(getattr(worksheet, "title", "")).replace("'", "''")
        start_letter = get_column_letter(start_column)
        end_letter = get_column_letter(end_column)
        metadata = spreadsheet.fetch_sheet_metadata(
            {
                "includeGridData": True,
                "ranges": [f"'{title}'!{start_letter}{source_row}:{end_letter}{source_row}"],
            }
        )
        worksheet_id = int(getattr(worksheet, "id"))
        sheet = next(
            item
            for item in metadata.get("sheets", [])
            if int(item.get("properties", {}).get("sheetId", -1)) == worksheet_id
        )
        grid = sheet.get("data", [{}])[0]
        source_cells = grid.get("rowData", [{}])[0].get("values", [])
        width = end_column - start_column + 1
        copied_cells: list[dict[str, Any]] = []
        for offset in range(width):
            source_cell = source_cells[offset] if offset < len(source_cells) else {}
            copied_cells.append(
                {
                    key: copy.deepcopy(source_cell[key])
                    for key in ("userEnteredFormat", "dataValidation", "note")
                    if key in source_cell
                }
            )

        requests: list[dict[str, Any]] = [
            {
                "updateCells": {
                    "range": {
                        "sheetId": worksheet_id,
                        "startRowIndex": target_row - 1,
                        "endRowIndex": target_row,
                        "startColumnIndex": start_column - 1,
                        "endColumnIndex": end_column,
                    },
                    "rows": [{"values": copied_cells}],
                    "fields": "userEnteredFormat,dataValidation,note",
                }
            }
        ]
        row_metadata = grid.get("rowMetadata", [])
        if row_metadata and "pixelSize" in row_metadata[0]:
            requests.append(
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": worksheet_id,
                            "dimension": "ROWS",
                            "startIndex": target_row - 1,
                            "endIndex": target_row,
                        },
                        "properties": {"pixelSize": row_metadata[0]["pixelSize"]},
                        "fields": "pixelSize",
                    }
                }
            )
        spreadsheet.batch_update({"requests": requests})
    except Exception as exc:
        logger.warning(
            "Could not copy income row properties",
            extra={
                "exception": str(exc),
                "source_row": source_row,
                "target_row": target_row,
            },
        )


def _repair_income_summary_formula(worksheet: Any, layout: dict[str, Any]) -> None:
    summary_row = worksheet.find("Monthly Income:").row
    first_income_row = layout["header_row"] + 1
    amount_column = layout["amount"]
    amount_letter = get_column_letter(amount_column)
    formula = f"=SUM({amount_letter}{first_income_row}:{amount_letter}{summary_row - 1})"
    _update_contiguous_row(worksheet, summary_row, [amount_column], [formula])


def _income_date_text(value: Any = None) -> str:
    if isinstance(value, datetime):
        parsed = value
    else:
        raw = str(value or "").strip()
        parsed = None
        if raw:
            for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    parsed = datetime.strptime(raw, pattern)
                    break
                except ValueError:
                    continue
            if parsed is None:
                return raw
        else:
            tz = ZoneInfo(os.getenv("TZ", "America/Los_Angeles"))
            parsed = datetime.now(tz)
    return f"{parsed.month}/{parsed.day}/{parsed.year}"


async def write_expense_to_sheet(data, message):
    category = normalize_expense_category(data.get("category"))
    if not category:
        if message:
            await message.channel.send("❌ Could not log entry — missing category.")
        return

    if category not in get_category_columns:
        available = ", ".join(sorted(get_category_columns.keys()))
        await message.channel.send(f"❌ Unknown category: {category}. Available categories: {available}.")
        return

    ws = None
    try:
        ws = await _expense_sheet_with_retry()
    except Exception as e:
        logger.exception("Could not access expense sheet", extra={"exception": str(e), "category": category})
        if message:
            await message.channel.send("❌ I could not access the expense sheet. Please try again in a moment.")
        return
    if ws is None:
        if message:
            await message.channel.send("❌ I could not access the expense sheet. Please try again in a moment.")
        return

    discord_user = getattr(message.author, "name", "").lower()
    discord_user_id = getattr(message.author, "id", None)
    discord_user_id = str(discord_user_id) if discord_user_id is not None else None

    # If the message explicitly mentions a non-bot user, treat that mention as the actor.
    # This lets people post via shortcuts/webhooks and still log under the correct account.
    for mentioned in getattr(message, "mentions", []) or []:
        if getattr(mentioned, "bot", False):
            continue
        discord_user = (getattr(mentioned, "name", None) or getattr(mentioned, "display_name", "")).lower()
        mentioned_id = getattr(mentioned, "id", discord_user_id)
        discord_user_id = str(mentioned_id) if mentioned_id is not None else None
        logger.debug("Using mentioned user for resolution", extra={"user": discord_user, "user_id": discord_user_id})
        break
    logger.debug("Discord user", extra={"user": discord_user})

    # Determine `person(s)` to log
    person = (data.get("person") or "").strip()
    if person.lower() in {"total", "all", "both", "everyone", "all persons", "all people"}:
        person = ""
    if not person:
        persons_to_log = resolve_query_persons(discord_user, None, discord_user_id)
        logger.debug("Resolved persons", extra={"persons": persons_to_log})
        if not persons_to_log:
            await message.channel.send("❌ Could not determine person for logging.")
            return
    else:
        persons_to_log = resolve_query_persons(discord_user, person, discord_user_id)
        logger.debug("Resolved explicit person", extra={"persons": persons_to_log})
        if not persons_to_log:
            await message.channel.send(f"❌ Could not resolve specified person: {person}")
            return

    # If multiple (ambiguous Brian), ask which card
    if len(persons_to_log) > 1:
        pending_data_by_user[discord_user] = {
            "data": data,
            "worksheet": ws,
            "category": category,
            "undo_user_key": get_current_discord_user_id() or discord_user_id,
        }

        async def handle_selection(interaction, selected_card):
            stored = pending_data_by_user.pop(discord_user, None)
            if not stored:
                await interaction.response.send_message("❌ Session expired.")
                return

            # Acknowledge quickly to avoid "Unknown interaction" errors, then send follow-up.
            try:
                await interaction.response.defer(ephemeral=True)
            except Exception:
                # If already acknowledged, continue to follow-up.
                pass

            stored["data"]["person"] = selected_card
            values = normalize_expense_data(stored["data"], selected_card)
            row = log_category_row(values, ws, category)
            record_expense_undo(
                category,
                row,
                values,
                selected_card,
                stored.get("undo_user_key"),
            )

            await interaction.followup.send(
                f"✅ {_logged_expense_label(category)} logged: ${stored['data']['amount']} for {selected_card}"
            )

        view = CardButtonView(handle_selection)
        await message.channel.send(
            f"{message.author.mention}, which card did you use?",
            view=view
        )
        return

    # Otherwise only one person
    selected_person = persons_to_log[0]
    data["person"] = selected_person
    values_to_write = normalize_expense_data(data, selected_person)

    # Validate required fields
    required_fields = ["amount", "person", "item"]
    missing = [f for f in required_fields if not values_to_write.get(f)]
    if missing:
        msg = f"❌ Could not log entry — missing: {', '.join(missing)}."
        print(msg)
        if message:
            await message.channel.send(msg)
        return

    row = log_category_row(values_to_write, ws, category)
    record_expense_undo(category, row, values_to_write, selected_person)

    if message:
        await message.channel.send(
            f"✅ {_logged_expense_label(category)} logged: ${data.get('amount')} for {selected_person}"
        )


def normalize_expense_data(data, person):
    tz = ZoneInfo(os.getenv("TZ", "America/Los_Angeles"))
    local_now = datetime.now(tz)
    return {
        "date": local_now.strftime("%-m/%-d/%Y"),
        "amount": float(data.get("amount") or 0),
        "location": (data.get("location") or "").strip(),
        "person": person.strip(),
        "item": (
            data.get("item") or
            data.get("food") or
            ""
        ).strip()
    }


def log_category_row(values, worksheet, category):
    logger.debug("Values passed to log_category_row", extra={"values": values})

    config = get_category_columns[category]
    row_start = config["start_row"]
    columns = config["columns"]

    ref_col_letter = columns.get("amount") or list(columns.values())[0]
    ref_col_index = column_index_from_string(ref_col_letter)
    col_values = worksheet.col_values(ref_col_index)[row_start - 1:]
    first_empty_row = len(col_values) + row_start

    write_columns = []
    write_values = []
    for field, col_letter in columns.items():
        value = values.get(field)
        logger.debug("Writing field", extra={"field": field, "column": col_letter, "value": value})
        if value is not None:
            col_index = column_index_from_string(col_letter)
            logger.debug("Writing cell", extra={"value": value, "row": first_empty_row, "col": col_index, "column": col_letter})
            write_columns.append(col_index)
            write_values.append(_sheet_user_entered_value(field, value))

    _update_contiguous_row(worksheet, first_empty_row, write_columns, write_values)

    logger.info("Logged expense row", extra={"category": category, "row": first_empty_row})
    return first_empty_row


def record_expense_undo(category, row, values_or_amount, person, user_key=None, metadata_extra: dict | None = None):
    columns = get_category_columns[category]["columns"]
    col_indexes = [column_index_from_string(col_letter) for col_letter in columns.values()]
    if isinstance(values_or_amount, dict):
        values = values_or_amount
        amount = values.get("amount")
        new_values = [str(values.get(field, "")) for field in columns.keys()]
    else:
        amount = values_or_amount
        new_values = []
    return record_undo_action(
        user_key or get_current_discord_user_id(),
        UndoAction(
            worksheet="expense",
            kind="clear_cells",
            row=row,
            columns=col_indexes,
            previous_values=["" for _ in col_indexes],
            new_values=new_values,
            metadata={"type": "expense", "category": category, "person": str(person), **(metadata_extra or {})},
            description=f"{category} expense ${amount} for {person}",
        ),
    )
