from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from datetime import date
import re
from typing import Any, Literal, Protocol

from openpyxl.utils import get_column_letter

from bookiebot.sheets.repo import get_sheets_repo
from bookiebot.sheets.routing import now_pacific
from bookiebot.sheets.utils import clean_money

BillRecurrence = Literal["monthly", "quarterly"]


class _WorksheetWithUpdate(Protocol):
    def update(self, *args: Any, **kwargs: Any) -> Any:
        ...

BILL_SCHEDULE_HEADERS = [
    "bill_key",
    "display_name",
    "recurrence",
    "pull_day",
    "pull_months",
    "source_label",
    "account",
    "notes",
    "updated_at",
]

DEFAULT_BILL_TEMPLATE_ROWS = [
    ("rent", "Rent", "monthly", "", "", "Rent", "", ""),
    ("pge", "PG&E", "monthly", "", "", "PG&E", "", ""),
    ("recology", "Recology", "quarterly", "", "", "Recology", "", ""),
    ("water", "Water", "quarterly", "", "", "Water", "", ""),
]

RETIRED_BILL_KEYS = {"student_loan", "student_loan_payment"}


@dataclass(frozen=True)
class BillSchedule:
    bill_key: str
    display_name: str
    recurrence: BillRecurrence
    pull_day: int
    pull_months: tuple[int, ...] = ()
    source_label: str = ""
    account: str = ""
    notes: str = ""
    source_range: str = ""
    updated_at: str = ""


@dataclass(frozen=True)
class BillReminder:
    bill: BillSchedule
    pull_date: date
    days_until: int
    amount: float | None
    amount_entered: bool
    overdue: bool = False

    @property
    def key(self) -> str:
        status = "overdue" if self.overdue else str(self.days_until)
        return f"{self.bill.bill_key}|{self.pull_date.isoformat()}|{status}"


@dataclass(frozen=True)
class BillScheduleWarning:
    source_range: str
    reason: str
    values: tuple[str, ...] = ()

    def format(self) -> str:
        detail = f" ({', '.join(value for value in self.values if value)})" if any(self.values) else ""
        return f"{self.source_range}: {self.reason}{detail}"


def _cell(row: list[str], index: int) -> str:
    if index >= len(row):
        return ""
    return str(row[index]).strip()


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _parse_day(value: str) -> int | None:
    text = value.strip().lower()
    if not text:
        return None
    if text in {"last", "last_day", "eom", "end_of_month"}:
        return 31
    match = re.search(r"\b([1-9]|[12][0-9]|3[01])(?:st|nd|rd|th)?\b", text)
    if not match:
        return None
    return int(match.group(1))


def _parse_pull_months(value: str) -> tuple[int, ...]:
    months = sorted({int(match) for match in re.findall(r"\d+", value)})
    return tuple(month for month in months if 1 <= month <= 12)


def _parse_recurrence(value: str) -> BillRecurrence | None:
    text = value.strip().lower()
    if text in {"monthly", "month"}:
        return "monthly"
    if text in {"quarterly", "quarter"}:
        return "quarterly"
    return None


def _source_range(row_index: int) -> str:
    return f"_BookieBot Bill Schedule!A{row_index}:I{row_index}"


def _row_is_empty(row: list[str]) -> bool:
    return not any(str(value).strip() for value in row)


def _row_is_blank_template(fields: dict[str, str]) -> bool:
    return bool(fields.get("display_name") or fields.get("source_label")) and not fields.get("pull_day", "").strip()


def _bill_from_fields(fields: dict[str, str], source_range: str) -> tuple[BillSchedule | None, BillScheduleWarning | None]:
    if not any(value.strip() for value in fields.values()):
        return None, None
    if _row_is_blank_template(fields):
        return None, None

    display_name = fields.get("display_name", "")
    source_label = fields.get("source_label", "") or display_name
    recurrence = _parse_recurrence(fields.get("recurrence", ""))
    pull_day = _parse_day(fields.get("pull_day", ""))
    pull_months = _parse_pull_months(fields.get("pull_months", ""))

    if not display_name:
        return None, BillScheduleWarning(source_range, "missing display name", tuple(fields.values()))
    if recurrence is None:
        return None, BillScheduleWarning(source_range, "invalid recurrence", (display_name, fields.get("recurrence", "")))
    if pull_day is None:
        return None, BillScheduleWarning(source_range, "missing or invalid pull day", (display_name, fields.get("pull_day", "")))
    if recurrence == "quarterly" and not pull_months:
        return None, BillScheduleWarning(source_range, "quarterly bill missing pull months", (display_name,))

    bill_key = fields.get("bill_key", "") or _slug(display_name)
    return (
        BillSchedule(
            bill_key=bill_key,
            display_name=display_name,
            recurrence=recurrence,
            pull_day=pull_day,
            pull_months=pull_months,
            source_label=source_label,
            account=fields.get("account", ""),
            notes=fields.get("notes", ""),
            source_range=source_range,
            updated_at=fields.get("updated_at", ""),
        ),
        None,
    )


def _update_range(ws: _WorksheetWithUpdate, start_row: int, start_col: int, values: list[list[str]]) -> None:
    if not values:
        return
    end_row = start_row + len(values) - 1
    end_col = start_col + max(len(row) for row in values) - 1
    range_name = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    try:
        ws.update(values, range_name=range_name, raw=False)
    except TypeError:
        ws.update(range_name, values, raw=False)


def _template_rows() -> list[list[str]]:
    updated_at = now_pacific().isoformat(timespec="seconds")
    return [list(row) + [updated_at] for row in DEFAULT_BILL_TEMPLATE_ROWS]


def ensure_bill_schedule_sheet() -> None:
    ws = get_sheets_repo().bill_schedule_sheet()
    rows = ws.get_all_values()
    if rows and any(any(str(value).strip() for value in row) for row in rows):
        return
    _update_range(ws, 1, 1, [BILL_SCHEDULE_HEADERS] + _template_rows())


def parse_bill_schedules_with_warnings(rows: list[list[str]] | None = None) -> tuple[list[BillSchedule], list[BillScheduleWarning]]:
    if rows is None:
        ensure_bill_schedule_sheet()
        rows = get_sheets_repo().bill_schedule_sheet().get_all_values()
    if not rows:
        return [], []

    header = [_normalize_header(value) for value in rows[0]]
    if not {"display_name", "recurrence", "pull_day", "source_label"} <= set(header):
        return [], [
            BillScheduleWarning(
                "_BookieBot Bill Schedule!A1:I1",
                "missing expected bill schedule headers",
                tuple(rows[0]),
            )
        ]

    bills: list[BillSchedule] = []
    warnings: list[BillScheduleWarning] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if _row_is_empty(row):
            continue
        fields = {
            column: _cell(row, index)
            for index, column in enumerate(header)
            if column
        }
        legacy_keys = {
            _slug(fields.get("bill_key", "")),
            _slug(fields.get("display_name", "")),
            _slug(fields.get("source_label", "")),
        }
        if legacy_keys & RETIRED_BILL_KEYS:
            continue
        bill, warning = _bill_from_fields(fields, _source_range(row_index))
        if bill:
            bills.append(bill)
        if warning:
            warnings.append(warning)
    return bills, warnings


def list_bill_schedules(rows: list[list[str]] | None = None) -> list[BillSchedule]:
    bills, _warnings = parse_bill_schedules_with_warnings(rows)
    return bills


def bill_amount_for_source_label(source_label: str) -> tuple[bool, float]:
    ws = get_sheets_repo().income_sheet()
    try:
        cell = ws.find(source_label)
        amount_cell = ws.cell(cell.row, cell.col + 1).value
    except Exception:
        return False, 0.0
    amount = clean_money(amount_cell or "")
    return amount > 0, amount


def _clamped_date(year: int, month: int, day: int) -> date:
    last_day = monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def _next_allowed_month(current_month: int, months: tuple[int, ...]) -> int:
    for month in months:
        if month >= current_month:
            return month
    return months[0]


def next_bill_pull_date(bill: BillSchedule, today: date) -> date | None:
    if bill.recurrence == "monthly":
        candidate = _clamped_date(today.year, today.month, bill.pull_day)
        if candidate < today:
            month = today.month + 1
            year = today.year
            if month == 13:
                month = 1
                year += 1
            candidate = _clamped_date(year, month, bill.pull_day)
        return candidate

    if not bill.pull_months:
        return None
    month = _next_allowed_month(today.month, bill.pull_months)
    year = today.year
    candidate = _clamped_date(year, month, bill.pull_day)
    if candidate < today:
        next_months = [allowed for allowed in bill.pull_months if allowed > today.month]
        if next_months:
            candidate = _clamped_date(today.year, next_months[0], bill.pull_day)
        else:
            candidate = _clamped_date(today.year + 1, bill.pull_months[0], bill.pull_day)
    return candidate


def overdue_bill_pull_date(bill: BillSchedule, today: date) -> date | None:
    if bill.recurrence == "monthly":
        candidate = _clamped_date(today.year, today.month, bill.pull_day)
    elif today.month in bill.pull_months:
        candidate = _clamped_date(today.year, today.month, bill.pull_day)
    else:
        return None
    return candidate if candidate < today else None


def due_bill_reminders_for_bills(bills: list[BillSchedule], today: date) -> list[BillReminder]:
    reminders: list[BillReminder] = []
    for bill in bills:
        amount_entered, amount = bill_amount_for_source_label(bill.source_label)
        pull_date = next_bill_pull_date(bill, today)
        if pull_date is not None:
            days_until = (pull_date - today).days
            if 0 <= days_until <= 7:
                reminders.append(BillReminder(bill, pull_date, days_until, amount if amount_entered else None, amount_entered))

        overdue_date = overdue_bill_pull_date(bill, today)
        if overdue_date is not None and not amount_entered:
            reminders.append(BillReminder(bill, overdue_date, (overdue_date - today).days, None, False, overdue=True))

    return sorted(reminders, key=lambda reminder: (reminder.overdue, reminder.pull_date, reminder.bill.display_name.lower()))


def due_bill_reminders(today: date) -> tuple[list[BillReminder], list[BillScheduleWarning]]:
    bills, warnings = parse_bill_schedules_with_warnings()
    return due_bill_reminders_for_bills(bills, today), warnings
