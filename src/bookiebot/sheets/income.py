from __future__ import annotations

import copy
import logging
from typing import Any

from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def income_sheet_layout(
    worksheet: Any,
    summary_row: int | None = None,
    *,
    rows: list[list[Any]] | None = None,
) -> dict[str, Any]:
    resolved_rows = rows if rows is not None else worksheet.get_all_values()
    resolved_summary_row = summary_row
    if resolved_summary_row is None:
        resolved_summary_row = next(
            (
                row_number
                for row_number, row in enumerate(resolved_rows, start=1)
                if any(str(value).strip() == "Monthly Income:" for value in row)
            ),
            None,
        )
    if resolved_summary_row is None:
        resolved_summary_row = int(worksheet.find("Monthly Income:").row)

    for row_number, row in enumerate(resolved_rows[: max(0, resolved_summary_row - 1)], start=1):
        columns_by_header: dict[str, int] = {}
        source_placeholder = "<Enter Source>"
        for column, value in enumerate(row, start=1):
            normalized = str(value).strip().rstrip(":").strip().lower()
            if normalized in {"date", "amount"}:
                columns_by_header[normalized] = column
            elif normalized in {"source", "employer"}:
                columns_by_header["source"] = column
                source_placeholder = "<Enter Employer>" if normalized == "employer" else "<Enter Source>"
        if "source" in columns_by_header and "amount" in columns_by_header:
            return {
                **columns_by_header,
                "header_row": row_number,
                "source_placeholder": source_placeholder,
            }

    return {
        "source": 2,
        "amount": 3,
        "header_row": 1,
        "source_placeholder": "<Enter Employer>",
    }


def repair_income_summary_formula(
    worksheet: Any,
    layout: dict[str, Any] | None = None,
    *,
    summary_row: int | None = None,
) -> None:
    resolved_summary_row = int(
        summary_row if summary_row is not None else worksheet.find("Monthly Income:").row
    )
    resolved_layout = layout or income_sheet_layout(worksheet, resolved_summary_row)
    first_income_row = resolved_layout["header_row"] + 1
    amount_column = resolved_layout["amount"]
    amount_letter = get_column_letter(amount_column)
    last_income_row = resolved_summary_row - 1
    formula = (
        f"=SUM({amount_letter}{first_income_row}:{amount_letter}{last_income_row})"
        if last_income_row >= first_income_row
        else "=0"
    )
    range_name = f"{amount_letter}{resolved_summary_row}"
    if hasattr(worksheet, "update"):
        try:
            worksheet.update([[formula]], range_name=range_name, raw=False)
        except TypeError:
            worksheet.update(range_name, [[formula]], raw=False)
        return
    worksheet.update_cell(resolved_summary_row, amount_column, formula)


def capture_income_row_properties(
    worksheet: Any,
    *,
    row: int,
    start_column: int,
    end_column: int,
) -> dict[str, Any] | None:
    spreadsheet = getattr(worksheet, "spreadsheet", None)
    if spreadsheet is None or not hasattr(spreadsheet, "fetch_sheet_metadata"):
        return None

    title = str(getattr(worksheet, "title", "")).replace("'", "''")
    start_letter = get_column_letter(start_column)
    end_letter = get_column_letter(end_column)
    metadata = spreadsheet.fetch_sheet_metadata(
        {
            "includeGridData": True,
            "ranges": [f"'{title}'!{start_letter}{row}:{end_letter}{row}"],
        }
    )
    worksheet_id = int(getattr(worksheet, "id"))
    sheet = next(
        item
        for item in metadata.get("sheets", [])
        if int(item.get("properties", {}).get("sheetId", -1)) == worksheet_id
    )
    grid = sheet.get("data", [{}])[0]
    source_cells = grid.get("rowData", [{}])[0].get("values", [])
    width = end_column - start_column + 1
    cells: list[dict[str, Any]] = []
    for offset in range(width):
        source_cell = source_cells[offset] if offset < len(source_cells) else {}
        cells.append(
            {
                key: copy.deepcopy(source_cell[key])
                for key in ("userEnteredFormat", "dataValidation", "note")
                if key in source_cell
            }
        )

    snapshot: dict[str, Any] = {"cells": cells}
    row_metadata = grid.get("rowMetadata", [])
    if row_metadata and "pixelSize" in row_metadata[0]:
        snapshot["pixelSize"] = row_metadata[0]["pixelSize"]
    return snapshot


def apply_income_row_properties(
    worksheet: Any,
    *,
    row: int,
    start_column: int,
    end_column: int,
    snapshot: dict[str, Any] | None,
) -> None:
    if snapshot is None:
        return
    spreadsheet = getattr(worksheet, "spreadsheet", None)
    if spreadsheet is None or not hasattr(spreadsheet, "batch_update"):
        return

    width = end_column - start_column + 1
    source_cells = snapshot.get("cells", [])
    cells = [
        copy.deepcopy(source_cells[offset]) if offset < len(source_cells) else {}
        for offset in range(width)
    ]
    worksheet_id = int(getattr(worksheet, "id"))
    requests: list[dict[str, Any]] = [
        {
            "updateCells": {
                "range": {
                    "sheetId": worksheet_id,
                    "startRowIndex": row - 1,
                    "endRowIndex": row,
                    "startColumnIndex": start_column - 1,
                    "endColumnIndex": end_column,
                },
                "rows": [{"values": cells}],
                "fields": "userEnteredFormat,dataValidation,note",
            }
        }
    ]
    if "pixelSize" in snapshot:
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": worksheet_id,
                        "dimension": "ROWS",
                        "startIndex": row - 1,
                        "endIndex": row,
                    },
                    "properties": {"pixelSize": snapshot["pixelSize"]},
                    "fields": "pixelSize",
                }
            }
        )
    spreadsheet.batch_update({"requests": requests})


def copy_income_row_properties(
    worksheet: Any,
    *,
    source_row: int,
    target_row: int,
    start_column: int,
    end_column: int,
) -> None:
    try:
        snapshot = capture_income_row_properties(
            worksheet,
            row=source_row,
            start_column=start_column,
            end_column=end_column,
        )
        apply_income_row_properties(
            worksheet,
            row=target_row,
            start_column=start_column,
            end_column=end_column,
            snapshot=snapshot,
        )
    except Exception as exc:
        logger.warning(
            "Could not copy income row properties",
            extra={
                "exception": str(exc),
                "source_row": source_row,
                "target_row": target_row,
            },
        )
