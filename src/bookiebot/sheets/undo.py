from __future__ import annotations

from dataclasses import asdict, dataclass, field
from datetime import datetime
import json
import logging
from typing import Any, Literal
import weakref
from uuid import uuid4

from openpyxl.utils import get_column_letter

from bookiebot.sheets.repo import get_sheets_repo
from bookiebot.sheets.routing import actor_key_aliases

logger = logging.getLogger(__name__)

WorksheetName = Literal["expense", "income"]
ActionKind = Literal["clear_cells", "delete_row", "restore_cells", "move_expense", "compact_category_cells"]


@dataclass
class UndoAction:
    worksheet: WorksheetName
    kind: ActionKind
    row: int
    columns: list[int]
    previous_values: list[str]
    description: str
    new_values: list[str] = field(default_factory=list)
    metadata: dict[str, str] = field(default_factory=dict)


_LAST_ACTION_BY_USER: dict[str, UndoAction] = {}
_GLOBAL_LAST_ACTION: UndoAction | None = None
_PENDING_DELETE_IDS_BY_USER: dict[str, list[str]] = {}
_PENDING_UPDATE_IDS_BY_USER: dict[str, list[str]] = {}
_PENDING_MOVE_IDS_BY_USER: dict[str, list[str]] = {}
_PENDING_UPDATE_FIELD_BY_USER: dict[str, tuple[str, str]] = {}
_PENDING_MOVE_ITEM_BY_USER: dict[str, tuple[str, str]] = {}
_RECENT_ACTION_OFFSET_BY_USER: dict[str, int] = {}
_LOG_HEADERS = ["id", "created_at", "user_key", "status", "undone_at", "action_json"]
_LOG_HEADER_READY: weakref.WeakSet[Any] = weakref.WeakSet()


def _sheet_value(value: Any) -> str:
    return "" if value is None else str(value)


def _currency_user_entered_value(value: Any) -> str:
    text = _sheet_value(value).strip()
    if not text or text.startswith("$"):
        return text
    try:
        amount = float(text.replace(",", ""))
    except ValueError:
        return text
    return f"${amount:.2f}"


def _money_display_value(value: Any) -> str:
    text = _sheet_value(value).strip()
    if not text:
        return ""
    try:
        amount = float(text.replace("$", "").replace(",", ""))
    except ValueError:
        return text if text.startswith("$") else f"${text}"
    return f"${amount:.2f}"


def _sheet_user_entered_value(field: str, value: Any) -> str:
    if field == "amount":
        return _currency_user_entered_value(value)
    return _sheet_value(value)


def _range_name(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    start = f"{get_column_letter(start_col)}{start_row}"
    end = f"{get_column_letter(end_col)}{end_row}"
    return start if start == end else f"{start}:{end}"


def _update_range(ws: Any, start_row: int, start_col: int, values: list[list[Any]]) -> None:
    if not values:
        return
    normalized = [[_sheet_value(value) for value in row] for row in values]
    max_width = max((len(row) for row in normalized), default=0)
    if max_width == 0:
        return
    padded = [row + [""] * (max_width - len(row)) for row in normalized]
    range_name = _range_name(start_row, start_col, start_row + len(padded) - 1, start_col + max_width - 1)
    if hasattr(ws, "update"):
        try:
            ws.update(padded, range_name=range_name, raw=False)
            return
        except TypeError:
            ws.update(range_name, padded, raw=False)
            return
    for row_offset, row_values in enumerate(padded):
        for col_offset, value in enumerate(row_values):
            ws.update_cell(start_row + row_offset, start_col + col_offset, value)


def _update_contiguous_row(ws: Any, row: int, columns: list[int], values: list[Any]) -> None:
    if not columns:
        return
    pairs = sorted(zip(columns, values, strict=False), key=lambda item: item[0])
    group_columns: list[int] = []
    group_values: list[Any] = []
    previous_col: int | None = None
    for col, value in pairs:
        if previous_col is not None and col != previous_col + 1:
            _update_range(ws, row, group_columns[0], [group_values])
            group_columns = []
            group_values = []
        group_columns.append(col)
        group_values.append(value)
        previous_col = col
    if group_columns:
        _update_range(ws, row, group_columns[0], [group_values])


@dataclass
class LoggedAction:
    id: str
    created_at: str
    user_key: str | None
    action: UndoAction
    status: Literal["active", "undone"] = "active"
    undone_at: str | None = None


@dataclass
class _LogRecord:
    row_index: int
    logged: LoggedAction


@dataclass
class _ActionLogData:
    ws: Any
    records: list[_LogRecord]


def _action_from_dict(payload: dict) -> UndoAction:
    return UndoAction(
        worksheet=payload["worksheet"],
        kind=payload["kind"],
        row=int(payload["row"]),
        columns=[int(col) for col in payload.get("columns", [])],
        previous_values=[_sheet_value(value) for value in payload.get("previous_values", [])],
        description=str(payload.get("description", "")),
        new_values=[_sheet_value(value) for value in payload.get("new_values", [])],
        metadata={str(k): str(v) for k, v in payload.get("metadata", {}).items()},
    )


def _log_sheet():
    return get_sheets_repo().action_log_sheet()


def _log_header_is_ready(ws: Any) -> bool:
    try:
        return ws in _LOG_HEADER_READY
    except TypeError:
        return bool(getattr(ws, "_bookiebot_log_header_ready", False))


def _mark_log_header_ready(ws: Any) -> None:
    try:
        _LOG_HEADER_READY.add(ws)
    except TypeError:
        try:
            setattr(ws, "_bookiebot_log_header_ready", True)
        except Exception:
            pass


def _ensure_log_header(ws: Any) -> None:
    if _log_header_is_ready(ws):
        return
    rows = ws.get_all_values()
    if rows and rows[0][: len(_LOG_HEADERS)] == _LOG_HEADERS:
        _mark_log_header_ready(ws)
        return
    _update_range(ws, 1, 1, [_LOG_HEADERS])
    _mark_log_header_ready(ws)


def _logged_action_from_row(row: list[str]) -> LoggedAction:
    padded = list(row) + [""] * (len(_LOG_HEADERS) - len(row))
    action_json = padded[5] or "{}"
    payload = json.loads(action_json)
    status: Literal["active", "undone"] = "undone" if padded[3] == "undone" else "active"
    return LoggedAction(
        id=padded[0],
        created_at=padded[1],
        user_key=padded[2] or None,
        status=status,
        undone_at=padded[4] or None,
        action=_action_from_dict(payload),
    )


def _read_log_data() -> _ActionLogData | None:
    try:
        ws = _log_sheet()
        _ensure_log_header(ws)
        rows = ws.get_all_values()
    except Exception:
        logger.exception("Failed to read action log worksheet")
        return None

    records: list[_LogRecord] = []
    for row_index, row in enumerate(rows[1:], start=2):
        if not row or not row[0]:
            continue
        try:
            records.append(_LogRecord(row_index=row_index, logged=_logged_action_from_row(row)))
        except Exception:
            logger.warning("Skipping malformed action log row", extra={"row": row})
    return _ActionLogData(ws=ws, records=records)


def _read_log() -> list[LoggedAction]:
    data = _read_log_data()
    if data is None:
        return []
    return [record.logged for record in data.records]


def read_active_logged_actions(user_key: str | None = None) -> list[LoggedAction]:
    keys = actor_key_aliases(str(user_key)) if user_key else set()
    return [
        logged
        for logged in _read_log()
        if logged.status == "active"
        and logged.action.metadata.get("type") != "system_state"
        and (not keys or logged.user_key in keys)
    ]


def _append_logged_action(user_key: str | None, action: UndoAction) -> None:
    ws = _log_sheet()
    _ensure_log_header(ws)
    logged = LoggedAction(
        id=uuid4().hex[:8],
        created_at=datetime.now().isoformat(timespec="seconds"),
        user_key=str(user_key) if user_key else None,
        action=action,
    )
    row = [
        logged.id,
        logged.created_at,
        logged.user_key or "",
        logged.status,
        logged.undone_at or "",
        json.dumps(asdict(logged.action), separators=(",", ":")),
    ]
    if hasattr(ws, "append_row"):
        ws.append_row(row)
        return
    rows = ws.get_all_values()
    ws.insert_row(row, index=len(rows) + 1)


def _find_log_row(logged_id: str, log_data: _ActionLogData | None = None) -> tuple[Any, int, LoggedAction] | None:
    data = log_data or _read_log_data()
    if data is None:
        return None
    for record in data.records:
        if record.logged.id == logged_id:
            return data.ws, record.row_index, record.logged
    return None


def _write_logged_action(ws: Any, row_index: int, logged: LoggedAction) -> None:
    _update_range(ws, row_index, 6, [[json.dumps(asdict(logged.action), separators=(",", ":"))]])


def _worksheet(name: WorksheetName):
    repo = get_sheets_repo()
    if name == "expense":
        return repo.expense_sheet()
    if name == "income":
        return repo.income_sheet()
    raise ValueError(f"Unsupported worksheet: {name}")


def record_undo_action(user_key: str | None, action: UndoAction) -> None:
    global _GLOBAL_LAST_ACTION
    _GLOBAL_LAST_ACTION = action
    if user_key:
        _LAST_ACTION_BY_USER[str(user_key)] = action
    try:
        _append_logged_action(user_key, action)
    except Exception:
        logger.exception("Failed to persist undo action")


def record_system_event(user_key: str | None, event_type: str, metadata: dict[str, str], description: str) -> bool:
    """Persist non-user-visible bot state in the action log."""
    payload = {"type": "system_state", "event_type": event_type, **metadata}
    try:
        _append_logged_action(
            user_key,
            UndoAction(
                worksheet="income",
                kind="restore_cells",
                row=0,
                columns=[],
                previous_values=[],
                new_values=[],
                metadata=payload,
                description=description,
            ),
        )
        return True
    except Exception:
        logger.exception("Failed to persist system event", extra={"event_type": event_type})
        return False


def has_system_event(user_key: str | None, event_type: str, metadata: dict[str, str]) -> bool:
    keys = actor_key_aliases(str(user_key)) if user_key else set()
    for logged in _read_log():
        action_metadata = logged.action.metadata
        if logged.status != "active":
            continue
        if action_metadata.get("type") != "system_state":
            continue
        if action_metadata.get("event_type") != event_type:
            continue
        if keys and logged.user_key not in keys:
            continue
        if all(action_metadata.get(key) == value for key, value in metadata.items()):
            return True
    return False


def _lineage_parent_id(action: UndoAction) -> str | None:
    return action.metadata.get("updated_action_id") or action.metadata.get("source_action_id") or None


def _lineage_id(logged: LoggedAction, actions_by_id: dict[str, LoggedAction]) -> str:
    current = logged
    seen = {logged.id}
    while parent_id := _lineage_parent_id(current.action):
        if parent_id in seen:
            break
        parent = actions_by_id.get(parent_id)
        if parent is None:
            break
        seen.add(parent_id)
        current = parent
    return current.id


def _dedupe_actions_by_lineage(actions: list[LoggedAction], all_actions: list[LoggedAction]) -> list[LoggedAction]:
    actions_by_id = {action.id: action for action in all_actions}
    seen_lineages: set[str] = set()
    deduped: list[LoggedAction] = []
    for action in reversed(actions):
        lineage_id = _lineage_id(action, actions_by_id)
        if lineage_id in seen_lineages:
            continue
        seen_lineages.add(lineage_id)
        deduped.append(action)
    return deduped


def recent_actions(user_key: str | None, limit: int = 5, offset: int = 0) -> list[LoggedAction]:
    keys = actor_key_aliases(str(user_key)) if user_key else set()
    data = _read_log_data()
    if data is None:
        return []
    all_actions = [record.logged for record in data.records]
    matches = [
        action
        for action in all_actions
        if action.status == "active"
        and action.action.metadata.get("type") not in {"delete", "system_state"}
        and (not keys or action.user_key in keys)
    ]
    matches = _dedupe_actions_by_lineage(matches, all_actions)
    start = max(offset, 0)
    end = start + max(limit, 1)
    return matches[start:end]


def _latest_raw_logged_action(
    user_key: str | None,
    log_data: _ActionLogData | None = None,
) -> LoggedAction | None:
    keys = actor_key_aliases(str(user_key)) if user_key else set()
    data = log_data or _read_log_data()
    if data is None:
        return None
    matches = [
        record.logged
        for record in data.records
        if record.logged.status == "active"
        and record.logged.action.metadata.get("type") != "system_state"
        and (not keys or record.logged.user_key in keys)
    ]
    return matches[-1] if matches else None


def _format_actions(actions: list[LoggedAction], *, empty_message: str, final_prompt: str) -> str:
    if not actions:
        return empty_message

    lines = ["Recent logged actions I can work with:"]
    for index, logged in enumerate(actions, start=1):
        title, data_lines = _format_action_list_item(index, logged.action)
        lines.append(title)
        lines.append("```")
        lines.extend(data_lines)
        lines.append("```")
    lines.append(final_prompt)
    return "\n".join(lines)


def format_recent_action_list(actions: list[LoggedAction], *, continued: bool = False) -> str:
    return _format_actions(
        actions,
        empty_message="I do not have any recent logged actions for you this month.",
        final_prompt="Type `show more` to continue." if continued else "Type `show more` to see older transactions.",
    )


def action_option_label(action: UndoAction) -> str:
    field_values = _field_values_for_action(action)
    category = action.metadata.get("category") or action.metadata.get("type") or "transaction"
    item = field_values.get("item")
    location = field_values.get("location")
    amount = field_values.get("amount")
    person = field_values.get("person") or action.metadata.get("person")
    amount_label = amount if str(amount).startswith("$") else f"${amount}" if amount else ""
    label_parts = [part for part in (item, location, amount_label, person) if part]
    label = " - ".join(label_parts) or action.description or category
    return label[:100]


def action_title(action: UndoAction) -> str:
    action_type = action.metadata.get("type")
    category = action.metadata.get("category")
    if action_type == "expense":
        return f"{category or 'expense'} expense".title()
    if action_type == "update":
        return f"Updated: {(category or 'transaction').capitalize()} Expense"
    if action_type == "move":
        source = action.metadata.get("source_category", "unknown")
        destination = action.metadata.get("destination_category") or category or "unknown"
        return f"Moved Expense: {source.capitalize()} -> {destination.capitalize()}"
    if action_type == "need_expense":
        return "Need Expense"
    if action_type == "payment":
        return f"{category or 'payment'} payment".title()
    if action_type == "savings":
        return f"{category or 'savings'} deposit".title()
    if action.metadata.get("source"):
        return "Income"
    return "Transaction"


def format_recent_actions(user_key: str | None, limit: int = 5, offset: int = 0) -> str:
    return _format_actions(
        recent_actions(user_key, limit, offset),
        empty_message="I do not have any recent logged actions for you this month.",
        final_prompt="Type `show more` to see older transactions.",
    )


def next_recent_actions_page(user_key: str | None, page_size: int = 5) -> tuple[str, list[LoggedAction]]:
    key = str(user_key) if user_key else ""
    offset = _RECENT_ACTION_OFFSET_BY_USER.get(key, 0)
    actions = recent_actions(user_key, page_size, offset)
    if actions and key:
        _RECENT_ACTION_OFFSET_BY_USER[key] = offset + page_size
    return (
        _format_actions(
            actions,
            empty_message="I do not have more recent logged actions for you this month.",
            final_prompt="Type `show more` to continue.",
        ),
        actions,
    )


def reset_recent_actions_page(user_key: str | None, next_offset: int = 5) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _RECENT_ACTION_OFFSET_BY_USER[key] = max(next_offset, 0)


def _action_search_text(logged: LoggedAction) -> str:
    action = logged.action
    searchable_metadata_keys = (
        "type",
        "category",
        "person",
        "source_category",
        "destination_category",
    )
    parts = [
        logged.id,
        action.description,
        action.worksheet,
        str(action.row),
        *action.new_values,
        *(action.metadata.get(key) for key in searchable_metadata_keys),
    ]
    return " ".join(str(part).lower() for part in parts if part is not None)


def select_recent_action(
    user_key: str | None,
    *,
    index: int | None = None,
    action_id: str | None = None,
    match_text: str | None = None,
    limit: int = 10,
) -> LoggedAction | None:
    actions = recent_actions(user_key, limit)
    if action_id:
        action_id = action_id.strip().lower()
        for logged in actions:
            if logged.id.lower() == action_id:
                return logged
    if index is not None and 1 <= index <= len(actions):
        return actions[index - 1]
    if match_text:
        needles = [part for part in match_text.lower().split() if part]
        for logged in actions:
            haystack = _action_search_text(logged)
            if all(needle in haystack for needle in needles):
                return logged
    return None


def matching_recent_actions(user_key: str | None, match_text: str, limit: int = 10) -> list[LoggedAction]:
    needles = [part for part in match_text.lower().split() if part]
    if not needles:
        return []
    matches = []
    for logged in recent_actions(user_key, limit):
        haystack = _action_search_text(logged)
        if all(needle in haystack for needle in needles):
            matches.append(logged)
    return matches


def format_delete_candidates(user_key: str | None, match_text: str, limit: int = 10) -> str:
    key = str(user_key) if user_key else ""
    actions = matching_recent_actions(user_key, match_text, limit)
    if key:
        _PENDING_DELETE_IDS_BY_USER[key] = [logged.id for logged in actions]
    return _format_actions(
        actions,
        empty_message=f"I could not find a recent logged action matching '{match_text}'.",
        final_prompt="Use the controls below, or type the number of the transaction you want to delete.",
    )


def format_update_candidates(user_key: str | None, match_text: str, limit: int = 10) -> str:
    key = str(user_key) if user_key else ""
    actions = matching_recent_actions(user_key, match_text, limit)
    if key:
        _PENDING_UPDATE_IDS_BY_USER[key] = [logged.id for logged in actions]
    return _format_actions(
        actions,
        empty_message=f"I could not find a recent logged action matching '{match_text}'.",
        final_prompt="Use the controls below, or type the number of the transaction you want to update.",
    )


def format_move_candidates(user_key: str | None, match_text: str, limit: int = 10) -> str:
    key = str(user_key) if user_key else ""
    actions = matching_recent_actions(user_key, match_text, limit)
    if key:
        _PENDING_MOVE_IDS_BY_USER[key] = [logged.id for logged in actions]
    return _format_actions(
        actions,
        empty_message=f"I could not find a recent logged action matching '{match_text}'.",
        final_prompt="Use the controls below, or type the number of the transaction you want to move followed by the new category.",
    )


def set_pending_update_selection(user_key: str | None, action_id: str) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_UPDATE_IDS_BY_USER[key] = [action_id]


def set_pending_delete_selection(user_key: str | None, action_id: str) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_DELETE_IDS_BY_USER[key] = [action_id]


def set_pending_move_selection(user_key: str | None, action_id: str) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_MOVE_IDS_BY_USER[key] = [action_id]


def clear_pending_action_selection(user_key: str | None) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_UPDATE_IDS_BY_USER.pop(key, None)
        _PENDING_DELETE_IDS_BY_USER.pop(key, None)
        _PENDING_MOVE_IDS_BY_USER.pop(key, None)
        _PENDING_UPDATE_FIELD_BY_USER.pop(key, None)
        _PENDING_MOVE_ITEM_BY_USER.pop(key, None)


def set_pending_update_field(user_key: str | None, action_id: str, field: str) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_UPDATE_FIELD_BY_USER[key] = (action_id, field)


def pending_update_field(user_key: str | None) -> tuple[str, str] | None:
    key = str(user_key) if user_key else ""
    return _PENDING_UPDATE_FIELD_BY_USER.get(key)


def set_pending_move_item(user_key: str | None, action_id: str, destination_category: str) -> None:
    key = str(user_key) if user_key else ""
    if key:
        _PENDING_MOVE_ITEM_BY_USER[key] = (action_id, destination_category)


def pending_move_item(user_key: str | None) -> tuple[str, str] | None:
    key = str(user_key) if user_key else ""
    return _PENDING_MOVE_ITEM_BY_USER.get(key)


def has_pending_action_selection(user_key: str | None) -> bool:
    key = str(user_key) if user_key else ""
    return bool(
        key
        and (
            _PENDING_UPDATE_IDS_BY_USER.get(key)
            or _PENDING_DELETE_IDS_BY_USER.get(key)
            or _PENDING_MOVE_IDS_BY_USER.get(key)
        )
    )


def pending_action_selection_kind(user_key: str | None) -> Literal["update", "delete", "move"] | None:
    key = str(user_key) if user_key else ""
    if key and _PENDING_UPDATE_IDS_BY_USER.get(key):
        return "update"
    if key and _PENDING_DELETE_IDS_BY_USER.get(key):
        return "delete"
    if key and _PENDING_MOVE_IDS_BY_USER.get(key):
        return "move"
    return None


def pending_action_selection_count(user_key: str | None, kind: Literal["update", "delete", "move"]) -> int:
    key = str(user_key) if user_key else ""
    if not key:
        return 0
    if kind == "update":
        return len(_PENDING_UPDATE_IDS_BY_USER.get(key, []))
    if kind == "delete":
        return len(_PENDING_DELETE_IDS_BY_USER.get(key, []))
    if kind == "move":
        return len(_PENDING_MOVE_IDS_BY_USER.get(key, []))
    return 0


def pending_action_selection_id(
    user_key: str | None,
    kind: Literal["update", "delete", "move"],
    index: int,
) -> str | None:
    key = str(user_key) if user_key else ""
    if not key or index < 1:
        return None
    if kind == "update":
        action_ids = _PENDING_UPDATE_IDS_BY_USER.get(key, [])
    elif kind == "delete":
        action_ids = _PENDING_DELETE_IDS_BY_USER.get(key, [])
    else:
        action_ids = _PENDING_MOVE_IDS_BY_USER.get(key, [])
    if index <= len(action_ids):
        return action_ids[index - 1]
    return None


def _field_columns_for_action(action: UndoAction) -> dict[str, int]:
    if action.worksheet == "expense":
        from bookiebot.sheets.config import get_category_columns
        from openpyxl.utils import column_index_from_string

        category = action.metadata.get("category")
        if category and category in get_category_columns:
            return {
                field: column_index_from_string(col_letter)
                for field, col_letter in get_category_columns[category]["columns"].items()
            }

    if action.metadata.get("type") in {"payment", "savings"}:
        return {"amount": action.columns[0]} if action.columns else {}

    return {}


def editable_fields_for_action(action: UndoAction) -> list[str]:
    fields = list(_field_columns_for_action(action).keys())
    return [field for field in fields if field != "date"]


def _field_values_for_action(action: UndoAction, values: list[str] | None = None) -> dict[str, str]:
    display_fields = action.metadata.get("display_fields")
    if display_fields:
        try:
            fields = [str(field) for field in json.loads(display_fields)]
        except Exception:
            fields = []
        source_values = list(values if values is not None else action.new_values)
        while len(source_values) < len(fields):
            source_values.append("")
        return {
            field: _sheet_value(source_values[index])
            for index, field in enumerate(fields)
        }

    field_columns = _field_columns_for_action(action)
    fields = list(field_columns.keys())
    source_values = list(values if values is not None else action.new_values)
    while len(source_values) < len(fields):
        source_values.append("")
    return {
        field: _sheet_value(source_values[index])
        for index, field in enumerate(fields)
    }


def _format_action_snapshot(action: UndoAction, values: list[str] | None = None) -> str:
    return "\n".join(_format_action_data_lines(action, values))


def format_action_detail_block(action: UndoAction) -> str:
    return "```\n" + _format_action_snapshot(action) + "\n```"


def _format_action_data_lines(action: UndoAction, values: list[str] | None = None) -> list[str]:
    field_values = _field_values_for_action(action, values)
    if action.metadata.get("type") in {"expense", "update", "move"}:
        return _field_data_lines(field_values, action)

    if action.metadata.get("type") == "income":
        return _income_data_lines(action, values)

    if action.metadata.get("type") in {"payment", "savings"}:
        amount = field_values.get("amount", "")
        if amount:
            return [f"   Amount: {_money_display_value(amount)}"]
        return [action.description]

    return [action.description]


def _format_action_list_item(index: int, action: UndoAction) -> tuple[str, list[str]]:
    field_values = _field_values_for_action(action)
    title = action_title(action)

    if action.metadata.get("type") == "income":
        lines = _income_data_lines(action)
    elif action.metadata.get("type") in {"payment", "savings"}:
        lines = _format_action_data_lines(action)
    else:
        lines = _field_data_lines(field_values, action)
    if not lines:
        lines = [action.description]
    return f"{index}. {title}", lines


def _income_data_lines(action: UndoAction, values: list[str] | None = None) -> list[str]:
    source_values = list(values if values is not None else action.new_values)
    while len(source_values) < 3:
        source_values.append("")
    description = _sheet_value(source_values[1]).strip()
    amount = _money_display_value(source_values[2])
    source = action.metadata.get("source") or description
    if amount and source:
        return [f"   Income: {amount} from {source}"]
    if amount:
        return [f"   Income: {amount}"]
    if description:
        return [f"   Income: {description}"]
    return [f"   {action.description}"]


def _field_data_lines(field_values: dict[str, str], action: UndoAction) -> list[str]:
    lines: list[str] = []
    display_fields = [
        ("date", "Date"),
        ("item", "Item"),
        ("location", "Location"),
        ("amount", "Amount"),
        ("person", "Person"),
    ]
    for field, label in display_fields:
        value = field_values.get(field)
        if not value and field == "person":
            value = action.metadata.get("person")
        if value:
            display_value = _money_display_value(value) if field == "amount" else value
            lines.append(f"   {label}: {display_value}")
    return lines


def _update_logged_new_values(
    logged_id: str,
    updates_by_col: dict[int, Any],
    log_data: _ActionLogData | None = None,
) -> None:
    found = _find_log_row(logged_id, log_data)
    if found is None:
        return
    ws, row_index, logged = found
    columns = logged.action.columns
    values = list(logged.action.new_values)
    while len(values) < len(columns):
        values.append("")
    for col, value in updates_by_col.items():
        if col in columns:
            values[columns.index(col)] = str(value)
    logged.action.new_values = values
    _write_logged_action(ws, row_index, logged)


def _category_columns(category: str) -> dict[str, int]:
    from openpyxl.utils import column_index_from_string

    from bookiebot.sheets.config import get_category_columns

    return {
        field: column_index_from_string(col_letter)
        for field, col_letter in get_category_columns[category]["columns"].items()
    }


def _first_empty_category_row(ws: Any, category: str) -> int:
    from bookiebot.sheets.config import get_category_columns

    config = get_category_columns[category]
    row_start = config["start_row"]
    columns = config["columns"]
    ref_col_letter = columns.get("amount") or list(columns.values())[0]
    ref_col_index = _category_columns(category)[next(field for field, col in columns.items() if col == ref_col_letter)]
    col_values = ws.col_values(ref_col_index)[row_start - 1:]
    for offset, value in enumerate(col_values):
        if not str(value).strip():
            return row_start + offset
    return len(col_values) + row_start


def _last_occupied_category_row(ws: Any, category: str) -> int:
    from bookiebot.sheets.config import get_category_columns

    config = get_category_columns[category]
    row_start = int(config["start_row"])
    columns = config["columns"]
    ref_col_letter = columns.get("amount") or list(columns.values())[0]
    ref_col_index = _category_columns(category)[next(field for field, col in columns.items() if col == ref_col_letter)]
    col_values = ws.col_values(ref_col_index)
    for row_index in range(len(col_values), row_start - 1, -1):
        if str(col_values[row_index - 1]).strip():
            return row_index
    return row_start - 1


def _category_snapshot(ws: Any, rows: range, columns: list[int]) -> list[list[str]]:
    return [
        [_sheet_value(ws.cell(row, col).value) for col in columns]
        for row in rows
    ]


def _restore_category_snapshot(ws: Any, start_row: int, columns: list[int], snapshot: list[list[str]]) -> None:
    if not snapshot or not columns:
        return
    _update_range(ws, start_row, min(columns), snapshot)


def _shift_category_cells_up(ws: Any, *, start_row: int, end_row: int, columns: list[int]) -> None:
    if not columns:
        return
    start_col = min(columns)
    if end_row < start_row:
        _update_range(ws, start_row, start_col, [[""] * len(columns)])
        return

    shifted_values = _category_snapshot(ws, range(start_row + 1, end_row + 1), columns)
    shifted_values.append([""] * len(columns))
    _update_range(ws, start_row, start_col, shifted_values)


def _action_category(action: UndoAction) -> str | None:
    if action.worksheet != "expense":
        return None
    return action.metadata.get("category")


def _shift_logged_action_rows(
    *,
    category: str,
    lower_row: int,
    upper_row: int,
    delta: int,
    exclude_ids: set[str] | None = None,
    log_data: _ActionLogData | None = None,
) -> None:
    if upper_row < lower_row:
        return
    exclude_ids = exclude_ids or set()
    data = log_data or _read_log_data()
    if data is None:
        raise RuntimeError("Could not read action log for row-reference updates.")
    for record in data.records:
        if record.logged.id in exclude_ids:
            continue
        logged = record.logged
        if logged.status != "active":
            continue

        changed = False
        if _action_category(logged.action) == category and lower_row <= logged.action.row <= upper_row:
            logged.action.row += delta
            changed = True

        source_category = logged.action.metadata.get("source_category")
        source_row = logged.action.metadata.get("source_row")
        if source_category == category and source_row:
            try:
                parsed_source_row = int(source_row)
            except ValueError:
                parsed_source_row = 0
            if lower_row <= parsed_source_row <= upper_row:
                logged.action.metadata["source_row"] = str(parsed_source_row + delta)
                changed = True

        if changed:
            _write_logged_action(data.ws, record.row_index, logged)


def _values_for_category(source_values: dict[str, str], destination_category: str, overrides: dict[str, Any]) -> dict[str, str]:
    from bookiebot.sheets.config import get_category_columns

    destination_fields = get_category_columns[destination_category]["columns"].keys()
    clean_overrides = {
        field: value
        for field, value in overrides.items()
        if value not in (None, "")
    }
    values = {
        field: _sheet_value(clean_overrides.get(field, source_values.get(field, "")))
        for field in destination_fields
    }
    return values


def _missing_required_move_fields(values: dict[str, str], destination_category: str) -> list[str]:
    required = ["date", "amount", "person"]
    if destination_category in {"food", "shopping"}:
        required.append("item")
    return [field for field in required if field in values and not str(values[field]).strip()]


def move_recent_action(
    user_key: str | None,
    *,
    destination_category: str | None,
    updates: dict[str, Any] | None = None,
    index: int | None = None,
    action_id: str | None = None,
    match_text: str | None = None,
) -> tuple[bool, str]:
    from bookiebot.sheets.config import get_category_columns

    updates = updates or {}
    destination_category = (destination_category or "").strip().lower()
    if match_text and not index and not action_id:
        return False, format_move_candidates(user_key, match_text, 10)

    key = str(user_key) if user_key else ""
    requested_index = index
    if index is not None and key and key in _PENDING_MOVE_IDS_BY_USER:
        candidate_ids = _PENDING_MOVE_IDS_BY_USER.get(key, [])
        if 1 <= index <= len(candidate_ids):
            action_id = candidate_ids[index - 1]
            index = None

    if not index and not action_id and not match_text:
        return False, format_recent_actions(user_key, 5)

    if destination_category not in get_category_columns:
        available = ", ".join(sorted(get_category_columns))
        return False, f"Tell me which category to move it to. Available categories: {available}."

    logged = select_recent_action(user_key, index=index, action_id=action_id, match_text=match_text)
    if logged is None and requested_index is not None and action_id:
        logged = select_recent_action(user_key, index=requested_index)
    if logged is None:
        return False, format_recent_actions(user_key, 5)

    action = logged.action
    if action.metadata.get("type") != "expense" or action.worksheet != "expense":
        return False, "I can only move normal expense rows between categories right now."

    source_category = action.metadata.get("category", "")
    if source_category == destination_category:
        return False, f"That expense is already in {destination_category}."

    source_values = _field_values_for_action(action)
    destination_values = _values_for_category(source_values, destination_category, updates)
    missing = _missing_required_move_fields(destination_values, destination_category)
    if missing:
        if missing == ["item"]:
            set_pending_move_item(user_key, logged.id, destination_category)
            return False, "What is the name of the item?"
        return False, f"I can move it to {destination_category}, but I still need: {', '.join(missing)}."

    ws = _worksheet("expense")
    source_columns = action.columns
    source_current_values = [_sheet_value(ws.cell(action.row, col).value) for col in source_columns]
    source_columns_by_field = _category_columns(source_category)
    source_category_columns = list(source_columns_by_field.values())
    source_end_row = max(action.row, _last_occupied_category_row(ws, source_category))
    source_snapshot = _category_snapshot(ws, range(action.row, source_end_row + 1), source_category_columns)
    destination_row = _first_empty_category_row(ws, destination_category)
    destination_columns_by_field = _category_columns(destination_category)
    destination_fields = list(destination_values.keys())
    destination_columns = [destination_columns_by_field[field] for field in destination_fields]
    destination_new_values = [destination_values[field] for field in destination_fields]
    destination_sheet_values = [
        _sheet_user_entered_value(field, destination_values[field])
        for field in destination_fields
    ]
    destination_previous_values = [_sheet_value(ws.cell(destination_row, col).value) for col in destination_columns]

    try:
        log_data = _read_log_data()
        if log_data is None:
            return False, "Something went wrong while reading the action log."
        _shift_category_cells_up(
            ws,
            start_row=action.row,
            end_row=source_end_row,
            columns=source_category_columns,
        )
        _update_contiguous_row(ws, destination_row, destination_columns, destination_sheet_values)
        _mark_undone(logged.id, log_data)
        _shift_logged_action_rows(
            category=source_category,
            lower_row=action.row + 1,
            upper_row=source_end_row,
            delta=-1,
            exclude_ids={logged.id},
            log_data=log_data,
        )
    except Exception as e:
        logger.exception("Failed to move recent action", extra={"exception": str(e)})
        return False, "Something went wrong while moving that expense."

    if key:
        _PENDING_MOVE_IDS_BY_USER.pop(key, None)
        _PENDING_MOVE_ITEM_BY_USER.pop(key, None)

    record_undo_action(
        user_key,
        UndoAction(
            worksheet="expense",
            kind="move_expense",
            row=destination_row,
            columns=destination_columns,
            previous_values=destination_previous_values,
            new_values=destination_new_values,
            metadata={
                "type": "move",
                "category": destination_category,
                "person": destination_values.get("person", ""),
                "source_action_id": logged.id,
                "source_category": source_category,
                "source_row": str(action.row),
                "source_columns": json.dumps(source_columns),
                "source_values": json.dumps(source_current_values),
                "source_category_columns": json.dumps(source_category_columns),
                "source_compact_start_row": str(action.row),
                "source_compact_end_row": str(source_end_row),
                "source_category_snapshot": json.dumps(source_snapshot),
                "destination_category": destination_category,
                "display_fields": json.dumps(destination_fields),
            },
            description=f"moved {source_category} expense to {destination_category}",
        ),
    )
    destination_display_action = UndoAction(
        worksheet="expense",
        kind="clear_cells",
        row=destination_row,
        columns=destination_columns,
        previous_values=[],
        new_values=destination_new_values,
        metadata={
            "type": "expense",
            "category": destination_category,
            "person": destination_values.get("person", ""),
            "display_fields": json.dumps(destination_fields),
        },
        description=f"{destination_category} expense",
    )
    return (
        True,
        "Moved logged expense:\n"
        "Before:\n"
        "```\n"
        f"{_format_action_snapshot(action, source_current_values)}\n"
        "```\n"
        "After:\n"
        "```\n"
        f"{_format_action_snapshot(destination_display_action, destination_new_values)}\n"
        "```",
    )


def update_recent_action(
    user_key: str | None,
    *,
    updates: dict[str, Any],
    index: int | None = None,
    action_id: str | None = None,
    match_text: str | None = None,
) -> tuple[bool, str]:
    normalized_updates = {
        str(field).strip().lower(): value
        for field, value in updates.items()
        if value not in (None, "")
    }
    if match_text and not normalized_updates and not index and not action_id:
        return False, format_update_candidates(user_key, match_text, 10)

    key = str(user_key) if user_key else ""
    requested_index = index
    if index is not None and key and key in _PENDING_UPDATE_IDS_BY_USER:
        candidate_ids = _PENDING_UPDATE_IDS_BY_USER.get(key, [])
        if 1 <= index <= len(candidate_ids):
            action_id = candidate_ids[index - 1]
            index = None

    logged = select_recent_action(
        user_key,
        index=index,
        action_id=action_id,
        match_text=match_text,
    )
    if logged is None and requested_index is not None and action_id:
        logged = select_recent_action(user_key, index=requested_index)
    if logged is None:
        return False, format_recent_actions(user_key, 5)

    field_columns = _field_columns_for_action(logged.action)
    if not field_columns:
        return False, "I found that action, but I do not know how to edit its fields yet."

    unknown = sorted(set(normalized_updates) - set(field_columns))
    if unknown:
        available = ", ".join(sorted(field_columns))
        return False, f"I can update {available} for that action, but not: {', '.join(unknown)}."

    if not normalized_updates:
        return False, f"I found {_format_action_snapshot(logged.action)}. Please specify the new value."

    ws = _worksheet(logged.action.worksheet)
    display_fields = list(field_columns.keys())
    before_values = list(logged.action.new_values)
    after_values = list(before_values)
    while len(after_values) < len(field_columns):
        after_values.append("")
    previous_values: list[str] = []
    columns: list[int] = []
    values: list[str] = []
    updates_by_col: dict[int, Any] = {}
    for field, value in normalized_updates.items():
        col = field_columns[field]
        previous_values.append(_sheet_value(ws.cell(logged.action.row, col).value))
        columns.append(col)
        values.append(_sheet_user_entered_value(field, value))
        updates_by_col[col] = value
        field_index = list(field_columns).index(field)
        after_values[field_index] = str(value)

    try:
        _update_contiguous_row(ws, logged.action.row, columns, values)
    except Exception as e:
        logger.exception("Failed to update recent action", extra={"exception": str(e)})
        return False, "Something went wrong while updating that logged action."

    _update_logged_new_values(logged.id, updates_by_col)
    if key:
        _PENDING_UPDATE_IDS_BY_USER.pop(key, None)
        _PENDING_UPDATE_FIELD_BY_USER.pop(key, None)
    record_undo_action(
        user_key,
        UndoAction(
            worksheet=logged.action.worksheet,
            kind="restore_cells",
            row=logged.action.row,
            columns=columns,
            previous_values=previous_values,
            new_values=after_values,
            metadata={
                **logged.action.metadata,
                "type": "update",
                "updated_action_id": logged.id,
                "display_fields": json.dumps(display_fields),
            },
            description=f"updated {logged.action.description}",
        ),
    )
    return (
        True,
        "Updated logged action:\n"
        "Before:\n"
        "```\n"
        f"{_format_action_snapshot(logged.action, before_values)}\n"
        "```\n"
        "After:\n"
        "```\n"
        f"{_format_action_snapshot(logged.action, after_values)}\n"
        "```",
    )


def _mark_undone(logged_id: str, log_data: _ActionLogData | None = None) -> None:
    found = _find_log_row(logged_id, log_data)
    if found is None:
        return
    ws, row_index, _logged = found
    _update_range(ws, row_index, 4, [["undone", datetime.now().isoformat(timespec="seconds")]])


def _mark_active(logged_id: str, log_data: _ActionLogData | None = None) -> None:
    found = _find_log_row(logged_id, log_data)
    if found is None:
        return
    ws, row_index, _logged = found
    _update_range(ws, row_index, 4, [["active", ""]])


def _latest_logged_action(
    user_key: str | None,
    log_data: _ActionLogData | None = None,
) -> LoggedAction | None:
    return _latest_raw_logged_action(user_key, log_data)


def _apply_undo_action(action: UndoAction, log_data: _ActionLogData | None = None) -> tuple[bool, str]:
    ws = _worksheet(action.worksheet)
    if action.kind == "move_expense":
        if "source_category_snapshot" in action.metadata:
            source_category = action.metadata["source_category"]
            source_start_row = int(action.metadata["source_compact_start_row"])
            source_end_row = int(action.metadata["source_compact_end_row"])
            source_columns = [int(col) for col in json.loads(action.metadata["source_category_columns"])]
            source_snapshot = json.loads(action.metadata["source_category_snapshot"])
            _shift_logged_action_rows(
                category=source_category,
                lower_row=source_start_row,
                upper_row=source_end_row - 1,
                delta=1,
                exclude_ids={action.metadata.get("source_action_id", "")},
                log_data=log_data,
            )
            _restore_category_snapshot(ws, source_start_row, source_columns, source_snapshot)
        else:
            source_row = int(action.metadata["source_row"])
            source_columns = [int(col) for col in json.loads(action.metadata["source_columns"])]
            source_values = [_sheet_value(value) for value in json.loads(action.metadata["source_values"])]
            _update_contiguous_row(ws, source_row, source_columns, source_values)
        _update_contiguous_row(ws, action.row, action.columns, action.previous_values)
    elif action.kind == "compact_category_cells":
        category = action.metadata["category"]
        snapshot = json.loads(action.metadata["category_snapshot"])
        start_row = int(action.metadata["compact_start_row"])
        end_row = int(action.metadata["compact_end_row"])
        _shift_logged_action_rows(
            category=category,
            lower_row=start_row,
            upper_row=end_row - 1,
            delta=1,
            exclude_ids={action.metadata.get("deleted_action_id", "")},
            log_data=log_data,
        )
        _restore_category_snapshot(ws, start_row, action.columns, snapshot)
    elif action.kind == "delete_row":
        if hasattr(ws, "delete_rows"):
            ws.delete_rows(action.row)
        elif hasattr(ws, "delete_row"):
            ws.delete_row(action.row)
        else:
            return False, "This sheet client cannot delete rows."
    elif action.kind == "clear_cells":
        _update_contiguous_row(ws, action.row, action.columns, [""] * len(action.columns))
    elif action.kind == "restore_cells":
        _update_contiguous_row(ws, action.row, action.columns, action.previous_values)
    else:
        return False, "Unknown undo action."
    return True, f"Undid: {action.description}"


def _delete_expense_action_with_compaction(user_key: str | None, logged: LoggedAction) -> tuple[bool, str]:
    action = logged.action
    category = _action_category(action)
    if action.metadata.get("type") not in {"expense", "update", "move"} or not category:
        return False, ""

    ws = _worksheet("expense")
    category_columns_by_field = _category_columns(category)
    columns = list(category_columns_by_field.values())
    end_row = max(action.row, _last_occupied_category_row(ws, category))
    rows = range(action.row, end_row + 1)
    snapshot = _category_snapshot(ws, rows, columns)

    try:
        log_data = _read_log_data()
        if log_data is None:
            return False, "Something went wrong while reading the action log."
        _shift_category_cells_up(ws, start_row=action.row, end_row=end_row, columns=columns)
        _mark_undone(logged.id, log_data)
        _shift_logged_action_rows(
            category=category,
            lower_row=action.row + 1,
            upper_row=end_row,
            delta=-1,
            exclude_ids={logged.id},
            log_data=log_data,
        )
    except Exception as e:
        logger.exception("Failed to compact deleted expense", extra={"exception": str(e)})
        return False, "Something went wrong while deleting that logged action."

    record_undo_action(
        user_key,
        UndoAction(
            worksheet="expense",
            kind="compact_category_cells",
            row=action.row,
            columns=columns,
            previous_values=[],
            new_values=[],
            metadata={
                "type": "delete",
                "category": category,
                "deleted_action_id": logged.id,
                "compact_start_row": str(action.row),
                "compact_end_row": str(end_row),
                "category_snapshot": json.dumps(snapshot),
            },
            description=f"deleted {action.description}",
        ),
    )
    return True, f"Deleted: {action.description}"


def delete_recent_action(
    user_key: str | None,
    *,
    index: int | None = None,
    action_id: str | None = None,
    match_text: str | None = None,
) -> tuple[bool, str]:
    key = str(user_key) if user_key else ""
    if match_text and not index and not action_id:
        return False, format_delete_candidates(user_key, match_text, 10)

    logged: LoggedAction | None = None
    requested_index = index
    if action_id:
        logged = select_recent_action(user_key, action_id=action_id)
    elif index is not None and key and key in _PENDING_DELETE_IDS_BY_USER:
        candidate_ids = _PENDING_DELETE_IDS_BY_USER.get(key, [])
        if 1 <= index <= len(candidate_ids):
            logged = select_recent_action(user_key, action_id=candidate_ids[index - 1])
    elif index is not None:
        logged = select_recent_action(user_key, index=index)
    if logged is None and requested_index is not None:
        logged = select_recent_action(user_key, index=requested_index)

    if logged is None:
        return False, format_recent_actions(user_key, 5)

    if logged.action.worksheet == "expense":
        success, detail = _delete_expense_action_with_compaction(user_key, logged)
        if detail:
            if success and key:
                _PENDING_DELETE_IDS_BY_USER.pop(key, None)
            return success, detail

    try:
        success, detail = _apply_undo_action(logged.action)
    except Exception as e:
        logger.exception("Failed to delete recent action", extra={"exception": str(e)})
        return False, "Something went wrong while deleting that logged action."

    if not success:
        return False, detail

    _mark_undone(logged.id)
    if key:
        _PENDING_DELETE_IDS_BY_USER.pop(key, None)
    return True, f"Deleted: {logged.action.description}"


def undo_last_action(user_key: str | None) -> tuple[bool, str]:
    global _GLOBAL_LAST_ACTION
    key = str(user_key) if user_key else None
    log_data = _read_log_data()
    if log_data is None:
        return False, "I could not read the action log right now. Please wait a minute and try undo again."
    logged = _latest_logged_action(key, log_data)
    if logged:
        action = logged.action
    elif key:
        action = _LAST_ACTION_BY_USER.pop(key, None)
        if action is None:
            return False, "I do not have a recent transaction for you to undo."
    else:
        action = _GLOBAL_LAST_ACTION
        if action is None:
            return False, "I do not have a recent transaction to undo."

    try:
        success, detail = _apply_undo_action(action, log_data)
    except Exception as e:
        logger.exception("Failed to undo last action", extra={"exception": str(e)})
        if key:
            _LAST_ACTION_BY_USER[key] = action
        return False, "Something went wrong while undoing the last transaction."
    if not success:
        return False, detail

    if logged:
        updated_action_id = action.metadata.get("updated_action_id")
        if updated_action_id and action.kind == "restore_cells":
            _update_logged_new_values(
                updated_action_id,
                {col: value for col, value in zip(action.columns, action.previous_values)},
                log_data,
            )
        deleted_action_id = action.metadata.get("deleted_action_id")
        if deleted_action_id and action.kind == "compact_category_cells":
            _mark_active(deleted_action_id, log_data)
        source_action_id = action.metadata.get("source_action_id")
        if source_action_id and action.kind == "move_expense":
            _mark_active(source_action_id, log_data)
        _mark_undone(logged.id, log_data)
    if _GLOBAL_LAST_ACTION is action:
        _GLOBAL_LAST_ACTION = None
    return True, detail
