from __future__ import annotations

from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
import re
from typing import Literal

from openpyxl.utils import get_column_letter

from bookiebot.sheets.repo import get_sheets_repo
from bookiebot.sheets.routing import get_current_discord_user_id, get_user_config, now_pacific
from bookiebot.sheets.utils import clean_money

SubscriptionCadence = Literal["monthly", "yearly"]

DEFAULT_REMINDER_OFFSETS = (7, 3, 1)
NORMALIZED_SCHEDULE_HEADERS = [
    "id",
    "active",
    "budget_owner_key",
    "owner_name",
    "kind",
    "cadence",
    "name",
    "amount",
    "pull_day",
    "pull_month",
    "account",
    "reminder_offsets",
    "source_range",
    "updated_at",
]


@dataclass(frozen=True)
class Subscription:
    name: str
    amount: float
    cadence: SubscriptionCadence
    pull_day: int | None = None
    pull_month: int | None = None
    id: str = ""
    budget_owner_key: str = ""
    owner_name: str = ""
    category: str = ""
    kind: str = ""
    account: str = ""
    active: bool = True
    reminder_offsets: tuple[int, ...] = DEFAULT_REMINDER_OFFSETS
    source_range: str = ""
    updated_at: str = ""


@dataclass(frozen=True)
class SubscriptionReminder:
    subscription: Subscription
    pull_date: date
    days_until: int

    @property
    def key(self) -> str:
        sub = self.subscription
        return f"{sub.name.lower()}|{sub.cadence}|{self.pull_date.isoformat()}|{self.days_until}"


@dataclass(frozen=True)
class SubscriptionParseWarning:
    source_range: str
    reason: str
    values: tuple[str, ...] = ()

    def format(self) -> str:
        detail = f" ({', '.join(value for value in self.values if value)})" if any(self.values) else ""
        return f"{self.source_range}: {self.reason}{detail}"


def _cell(row: list[str], index: int) -> str:
    if index >= len(row):
        return ""
    return str(row[index]).strip()


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.strip().lower()).strip("_")


def _parse_day(value: str) -> int | None:
    text = value.strip().lower()
    if not text:
        return None
    if text in {"last", "last_day", "eom", "end_of_month"}:
        return 31
    match = re.search(r"\b([1-9]|[12][0-9]|3[01])(?:st|nd|rd|th)?\b", text)
    if not match:
        return None
    return int(match.group(1))


def _parse_month_day(value: str) -> tuple[int | None, int | None]:
    text = value.strip()
    if not text:
        return None, None
    match = re.search(r"\b(1[0-2]|0?[1-9])\s*/\s*([12]?[0-9]|3[01])(?:\s*/\s*\d{2,4})?\b", text)
    if match:
        return int(match.group(1)), int(match.group(2))

    try:
        parsed = datetime.strptime(text, "%B %d")
        return parsed.month, parsed.day
    except ValueError:
        pass

    try:
        parsed = datetime.strptime(text, "%b %d")
        return parsed.month, parsed.day
    except ValueError:
        return None, None


def _parse_offsets(value: str) -> tuple[int, ...]:
    if not value.strip():
        return DEFAULT_REMINDER_OFFSETS
    offsets = sorted({int(match) for match in re.findall(r"\d+", value)})
    return tuple(offset for offset in offsets if offset >= 0) or DEFAULT_REMINDER_OFFSETS


def _format_offsets(offsets: tuple[int, ...]) -> str:
    return ",".join(str(offset) for offset in offsets)


def _parse_active(value: str) -> bool:
    text = value.strip().lower()
    return text not in {"no", "n", "false", "inactive", "cancelled", "canceled", "0"}


def _parse_optional_int(value: str) -> int | None:
    text = value.strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        return None


def _parse_cadence(value: str, fallback: SubscriptionCadence | None = None) -> SubscriptionCadence | None:
    text = value.strip().lower()
    if "year" in text or "annual" in text:
        return "yearly"
    if "month" in text:
        return "monthly"
    return fallback


def _subscription_id(owner_key: str, name: str, cadence: str, source_range: str = "") -> str:
    base = re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_") or "subscription"
    range_part = re.sub(r"[^a-z0-9]+", "_", source_range.strip().lower()).strip("_")
    suffix = f":{range_part}" if range_part else ""
    return f"{owner_key or 'unknown'}:{base}:{cadence or 'unknown'}{suffix}"


def _subscription_from_fields(
    fields: dict[str, str],
    fallback_cadence: SubscriptionCadence | None = None,
    *,
    source_range: str = "",
    owner_key: str = "",
    owner_name: str = "",
    updated_at: str = "",
) -> Subscription | None:
    name = fields.get("name") or fields.get("service") or fields.get("merchant") or ""
    amount_text = fields.get("amount") or fields.get("estimate") or fields.get("expected_amount") or ""
    if not name.strip() or not amount_text.strip():
        return None

    cadence = _parse_cadence(fields.get("cadence", "") or fields.get("frequency", ""), fallback_cadence)
    if cadence is None:
        cadence = "yearly" if (fields.get("pull_date") or fields.get("date")) else "monthly"

    pull_day = _parse_day(fields.get("pull_day", "") or fields.get("day", "") or fields.get("recurring", ""))
    pull_month = None
    if cadence == "yearly":
        pull_date = fields.get("pull_date", "") or fields.get("date", "") or fields.get("recurring", "")
        pull_month, pull_day_from_date = _parse_month_day(pull_date)
        pull_day = pull_day_from_date or pull_day

    pull_day = _parse_optional_int(fields.get("pull_day", "")) or pull_day
    pull_month = _parse_optional_int(fields.get("pull_month", "")) or pull_month
    if pull_day is None:
        return None

    budget_owner_key = fields.get("budget_owner_key", "") or owner_key
    resolved_source_range = fields.get("source_range", "") or source_range
    resolved_updated_at = fields.get("updated_at", "") or updated_at
    resolved_owner_name = fields.get("owner_name", "") or owner_name
    active_value = fields.get("active", "yes") or fields.get("status", "active")
    subscription_id = fields.get("id", "") or _subscription_id(budget_owner_key, name, cadence, resolved_source_range)

    return Subscription(
        name=name.strip(),
        amount=clean_money(amount_text),
        cadence=cadence,
        pull_day=pull_day,
        pull_month=pull_month,
        id=subscription_id,
        budget_owner_key=budget_owner_key,
        owner_name=resolved_owner_name,
        category=(fields.get("category") or "").strip(),
        kind=(fields.get("kind") or fields.get("type") or "").strip(),
        account=(fields.get("account") or fields.get("card") or "").strip(),
        active=_parse_active(active_value),
        reminder_offsets=_parse_offsets(fields.get("reminder_offsets", "") or fields.get("reminders", "")),
        source_range=resolved_source_range,
        updated_at=resolved_updated_at,
    )


def _parse_warning_for_fields(
    fields: dict[str, str],
    cadence: SubscriptionCadence,
    source_range: str,
) -> SubscriptionParseWarning | None:
    schedule = fields.get("recurring", "") or fields.get("pull_day", "") or fields.get("pull_date", "") or fields.get("date", "")
    name = fields.get("name", "") or fields.get("service", "") or fields.get("merchant", "")
    amount = fields.get("amount", "") or fields.get("estimate", "") or fields.get("expected_amount", "")
    if not name.strip():
        return SubscriptionParseWarning(source_range, "missing subscription name", (schedule, amount))
    if not amount.strip():
        return SubscriptionParseWarning(source_range, "missing amount", (schedule, name))
    if not schedule.strip():
        return SubscriptionParseWarning(source_range, "missing pull date", (name, amount))
    if cadence == "yearly":
        month, day = _parse_month_day(schedule)
        if month is None or day is None:
            return SubscriptionParseWarning(source_range, f'invalid yearly date "{schedule}"', (name, amount))
    elif _parse_day(schedule) is None:
        return SubscriptionParseWarning(source_range, f'invalid monthly day "{schedule}"', (name, amount))
    return None


def _parse_normalized_table(rows: list[list[str]]) -> list[Subscription]:
    if not rows:
        return []
    for header_index, row in enumerate(rows[:10]):
        normalized = [_normalize_header(value) for value in row]
        if "name" not in normalized and "service" not in normalized:
            continue
        if "amount" not in normalized:
            continue
        if normalized.count("name") + normalized.count("service") > 1 or normalized.count("amount") > 1:
            continue
        if not ({"pull_day", "day"} & set(normalized) or {"pull_date", "date"} & set(normalized)):
            continue

        subscriptions: list[Subscription] = []
        for data_row in rows[header_index + 1 :]:
            fields = {
                header: _cell(data_row, index)
                for index, header in enumerate(normalized)
                if header
            }
            subscription = _subscription_from_fields(fields)
            if subscription:
                subscriptions.append(subscription)
        return subscriptions
    return []


def _current_owner_metadata() -> tuple[str, str]:
    actor_key = get_current_discord_user_id()
    user_config = get_user_config(actor_key)
    return user_config.budget_owner_key, user_config.name


def _block_metadata(rows: list[list[str]], header_row_index: int, start_col: int) -> tuple[str, SubscriptionCadence]:
    top = " ".join(
        _cell(rows[row_index], col_index)
        for row_index in range(max(0, header_row_index - 3), header_row_index)
        for col_index in range(start_col, min(start_col + 4, len(rows[row_index])))
    ).lower()
    kind = "wants" if "want" in top else "needs" if "need" in top else ""
    cadence: SubscriptionCadence = "yearly" if "year" in top else "monthly"
    return kind, cadence


def _parse_block_layout(
    rows: list[list[str]],
    *,
    owner_key: str = "",
    owner_name: str = "",
    updated_at: str = "",
    warnings: list[SubscriptionParseWarning] | None = None,
) -> list[Subscription]:
    subscriptions: list[Subscription] = []
    for row_index, row in enumerate(rows):
        normalized = [_normalize_header(value) for value in row]
        for start_col, header in enumerate(normalized):
            if header not in {"recurring", "date"}:
                continue
            next_headers = normalized[start_col : start_col + 3]
            if "name" not in next_headers or "amount" not in next_headers:
                continue

            kind, cadence = _block_metadata(rows, row_index, start_col)
            if header == "date":
                cadence = "yearly" if cadence == "yearly" else cadence

            for data_row_index, data_row in enumerate(rows[row_index + 1 :], start=row_index + 2):
                schedule = _cell(data_row, start_col)
                name = _cell(data_row, start_col + 1)
                amount = _cell(data_row, start_col + 2)
                if not schedule and not name and not amount:
                    break
                source_range = (
                    f"Subscriptions!"
                    f"{get_column_letter(start_col + 1)}{data_row_index}:"
                    f"{get_column_letter(start_col + 3)}{data_row_index}"
                )
                fields = {
                    "name": name,
                    "amount": amount,
                    "kind": kind,
                    "recurring": schedule,
                }
                subscription = _subscription_from_fields(
                    fields,
                    fallback_cadence=cadence,
                    source_range=source_range,
                    owner_key=owner_key,
                    owner_name=owner_name,
                    updated_at=updated_at,
                )
                if subscription:
                    subscriptions.append(subscription)
                elif warnings is not None:
                    warning = _parse_warning_for_fields(fields, cadence, source_range)
                    if warning:
                        warnings.append(warning)
    return subscriptions


def list_subscription_schedules(rows: list[list[str]] | None = None) -> list[Subscription]:
    if rows is None:
        rows = get_sheets_repo().subscriptions_sheet().get_all_values()
    subscriptions = _parse_normalized_table(rows)
    if subscriptions:
        return [subscription for subscription in subscriptions if subscription.active]
    return [subscription for subscription in _parse_block_layout(rows) if subscription.active]


def parse_visible_subscription_schedules(rows: list[list[str]] | None = None) -> list[Subscription]:
    subscriptions, _warnings = parse_visible_subscription_schedules_with_warnings(rows)
    return subscriptions


def parse_visible_subscription_schedules_with_warnings(
    rows: list[list[str]] | None = None,
) -> tuple[list[Subscription], list[SubscriptionParseWarning]]:
    if rows is None:
        rows = get_sheets_repo().subscriptions_sheet().get_all_values()

    updated_at = now_pacific().isoformat(timespec="seconds")
    try:
        owner_key, owner_name = _current_owner_metadata()
    except Exception:
        owner_key, owner_name = "", ""

    warnings: list[SubscriptionParseWarning] = []
    subscriptions = _parse_normalized_table(rows)
    if not subscriptions:
        subscriptions = _parse_block_layout(
            rows,
            owner_key=owner_key,
            owner_name=owner_name,
            updated_at=updated_at,
            warnings=warnings,
        )

    normalized: list[Subscription] = []
    for subscription in subscriptions:
        normalized.append(
            Subscription(
                id=subscription.id or _subscription_id(owner_key, subscription.name, subscription.cadence, subscription.source_range),
                active=subscription.active,
                budget_owner_key=subscription.budget_owner_key or owner_key,
                owner_name=subscription.owner_name or owner_name,
                kind=subscription.kind,
                cadence=subscription.cadence,
                name=subscription.name,
                amount=subscription.amount,
                pull_day=subscription.pull_day,
                pull_month=subscription.pull_month,
                account=subscription.account,
                reminder_offsets=subscription.reminder_offsets,
                source_range=subscription.source_range,
                updated_at=subscription.updated_at or updated_at,
                category=subscription.category,
            )
        )
    return normalized, warnings


def _subscription_to_row(subscription: Subscription) -> list[str]:
    return [
        subscription.id,
        "yes" if subscription.active else "no",
        subscription.budget_owner_key,
        subscription.owner_name,
        subscription.kind,
        subscription.cadence,
        subscription.name,
        f"{subscription.amount:.2f}",
        str(subscription.pull_day or ""),
        str(subscription.pull_month or ""),
        subscription.account,
        _format_offsets(subscription.reminder_offsets),
        subscription.source_range,
        subscription.updated_at,
    ]


def _update_range(ws: object, start_row: int, start_col: int, values: list[list[str]]) -> None:
    if not values:
        return
    end_row = start_row + len(values) - 1
    end_col = start_col + max(len(row) for row in values) - 1
    range_name = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
    try:
        ws.update(values, range_name=range_name, raw=False)
    except TypeError:
        ws.update(range_name, values, raw=False)


def _write_subscription_schedule_rows(subscriptions: list[Subscription]) -> None:
    rows = [NORMALIZED_SCHEDULE_HEADERS] + [_subscription_to_row(subscription) for subscription in subscriptions]
    ws = get_sheets_repo().subscription_schedule_sheet()
    existing_rows = ws.get_all_values()
    rows_to_write = max(len(rows), len(existing_rows), 1)
    width = len(NORMALIZED_SCHEDULE_HEADERS)
    padded_rows = [
        (rows[index] if index < len(rows) else [""] * width)
        for index in range(rows_to_write)
    ]
    padded_rows = [row + [""] * (width - len(row)) for row in padded_rows]
    _update_range(ws, 1, 1, padded_rows)


def sync_subscription_schedule_sheet() -> list[Subscription]:
    subscriptions = parse_visible_subscription_schedules()
    _write_subscription_schedule_rows(subscriptions)
    return subscriptions


def debug_subscription_sync() -> tuple[list[Subscription], list[SubscriptionParseWarning]]:
    subscriptions, warnings = parse_visible_subscription_schedules_with_warnings()
    _write_subscription_schedule_rows(subscriptions)
    return subscriptions, warnings


def list_normalized_subscription_schedules() -> list[Subscription]:
    rows = get_sheets_repo().subscription_schedule_sheet().get_all_values()
    if not rows:
        return []
    subscriptions = _parse_normalized_table(rows)
    return [subscription for subscription in subscriptions if subscription.active]


def _clamped_date(year: int, month: int, day: int) -> date:
    last_day = monthrange(year, month)[1]
    return date(year, month, min(day, last_day))


def next_pull_date(subscription: Subscription, today: date) -> date | None:
    if subscription.pull_day is None:
        return None

    if subscription.cadence == "monthly":
        candidate = _clamped_date(today.year, today.month, subscription.pull_day)
        if candidate < today:
            next_month = today.month + 1
            next_year = today.year
            if next_month == 13:
                next_month = 1
                next_year += 1
            candidate = _clamped_date(next_year, next_month, subscription.pull_day)
        return candidate

    if subscription.pull_month is None:
        return None
    candidate = _clamped_date(today.year, subscription.pull_month, subscription.pull_day)
    if candidate < today:
        candidate = _clamped_date(today.year + 1, subscription.pull_month, subscription.pull_day)
    return candidate


def due_subscription_reminders_for_subscriptions(
    subscriptions: list[Subscription],
    today: date,
) -> list[SubscriptionReminder]:
    reminders: list[SubscriptionReminder] = []
    for subscription in subscriptions:
        if not subscription.active:
            continue
        pull_date = next_pull_date(subscription, today)
        if pull_date is None:
            continue
        days_until = (pull_date - today).days
        if days_until in subscription.reminder_offsets:
            reminders.append(SubscriptionReminder(subscription, pull_date, days_until))
    return sorted(reminders, key=lambda reminder: (reminder.pull_date, reminder.subscription.name.lower()))


def due_subscription_reminders(today: date | None = None) -> list[SubscriptionReminder]:
    current = today or date.today()
    try:
        subscriptions = sync_subscription_schedule_sheet()
    except Exception:
        subscriptions = []

    if not subscriptions:
        try:
            subscriptions = list_normalized_subscription_schedules()
        except Exception:
            subscriptions = list_subscription_schedules()

    return due_subscription_reminders_for_subscriptions(subscriptions, current)


def format_subscription_reminder(reminder: SubscriptionReminder, mention: str | None = None) -> str:
    subscription = reminder.subscription
    when = "tomorrow" if reminder.days_until == 1 else f"in {reminder.days_until} days"
    prefix = f"{mention} " if mention else ""
    amount = f"${subscription.amount:.2f}" if subscription.amount else "an unknown amount"
    account = f" from {subscription.account}" if subscription.account else ""
    pull_date = f"{reminder.pull_date:%b} {reminder.pull_date.day}"
    return (
        f"{prefix}Reminder: {subscription.name} is expected to pull {amount}{account} "
        f"{when} ({pull_date})."
    )
